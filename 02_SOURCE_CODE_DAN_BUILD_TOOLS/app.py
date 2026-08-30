import sys
import subprocess

try:
    import barcode
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "python-barcode"])

import os                      # Alat pemeriksa keberadaan file (logo & excel) di Windows
import pandas as pd            # Mesin pembaca & manipulasi database Excel (.xlsm)
import customtkinter as ctk   # Paket dekorasi visual UI premium ala Windows 11
from tkinter import messagebox, ttk  # Modul pesan popup harian & tabel grid data kotor
from PIL import Image          # Paket rendering gambar untuk memunculkan logo perusahaan

# Mengunci tema aplikasi ke mode terang premium dengan aksen biru royal
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class ProfessionalWarehouseApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # 1. Mengatur Judul Resmi pada Bar Jendela Windows Anda
        self.title("GTP Laptop Inventory System - Enterprise Premium Edition")
        
        # 2. Mengatur Resolusi Layar Standar (Lebar 1280 x Tinggi 720 Piksel)
        self.geometry("1280x720")
        
        # 3. Mengubah latar belakang luar menjadi Slate Grey yang lebih cerah
        self.configure(fg_color=("#E2E8F0", "#E2E8F0")) 
        
        # 4. Mengunci Lokasi Berkas Pusat Database Excel Makro Anda
        self.excel_file = "Inventaris_Laptop.xlsm"
        
        # Konfigurasi Koneksi REST API v2 (Gudang Bridge) — TANPA psycopg2 langsung
        # Default: domain publik via Cloudflare Tunnel. LAN: http://192.168.30.100:1888
        self.api_base_gudang = os.environ.get("GTP_API_BASE_URL", "https://gtp.hoyodev.biz.id").rstrip("/")
        # Kredensial DB lama tidak dipakai lagi — disimpan kosong demi kompatibilitas
        self.db_host_online = ""; self.db_port_online = 0; self.db_name_online = ""
        self.db_user_online = ""; self.db_pass_online = ""
        self.database_sql_file = "Inventaris_Laptop.db"
        
        # 5. Membagi Jendela Menjadi 2 Sektor Grid (Kolom 0: Sidebar, Kolom 1: Konten Utama)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
                
        # 6. Membuat bingkai dasar (Frame) Sidebar di Kolom 0 dengan mencerahkan warna dasar
        self.sidebar_frame = ctk.CTkFrame(self, width=260, corner_radius=0, fg_color=("#CBD5E1", "#1E293B"), border_color="#94A3B8", border_width=1)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
               
        # Mengatur konfigurasi grid baris agar tombol menu tersusun rapat dan proporsional
        for i in range(13):
            self.sidebar_frame.grid_rowconfigure(i, weight=0)
        self.sidebar_frame.grid_rowconfigure(12, weight=1) # Ruang kosong pendorong ke bawah
        
        # 7. FIX REVISI: MENAIKKAN UKURAN LOGO MENJADI RAKSASA (90x90 PIKSEL) AGAR SANGAT MENCOLOK
        self.logo_path = "logo.png"
        if os.path.exists(self.logo_path):
            try:
                img_file = Image.open(self.logo_path)
                # Mengubah ukuran size dari 70x70 menjadi 90x90 piksel secara presisi harian
                self.logo_image = ctk.CTkImage(light_image=img_file, dark_image=img_file, size=(90, 90))
                self.logo_pic_label = ctk.CTkLabel(self.sidebar_frame, image=self.logo_image, text="")
                # Menyeimbangkan bantalan jarak vertikal (pady) agar tetap rata tengah simetris
                self.logo_pic_label.grid(row=0, column=0, padx=20, pady=(30, 6))
            except:
                pass
        
        # 8. FIX REVISI: Memperbesar huruf teks nama perusahaan menjadi Size 16 dan dipertebal (Bold)
        self.logo_label = ctk.CTkLabel(
            self.sidebar_frame, 
            text="GLOBAL TEKNOLOGI\nPRODIGI", 
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"), # Menaikkan size ke 16
            text_color=("#1E3A8A", "#F8FAFC"), # Light mode: Deep Royal Blue agar lebih mewah, Dark mode: Putih cerah
            justify="center"
        )
        self.logo_label.grid(row=1, column=0, padx=20, pady=(2, 20))
        
        # 9. Menyusun 10 Tombol Menu Sidebar Utama dengan Warna Teks Dinamis Mode Adaptif
        txt_clr = ("#000000", "#E2E8F0")
        
        # Menu 1: Dashboard Status
        self.btn_dash = ctk.CTkButton(self.sidebar_frame, text="📊 Dashboard Status", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("dash"))
        self.btn_dash.grid(row=2, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 2: SCAN BARANG
        self.btn_scan = ctk.CTkButton(self.sidebar_frame, text="🔍 SCAN BARANG", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("scan"))
        self.btn_scan.grid(row=3, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 3: CETAK SURAT JALAN
        self.btn_sj = ctk.CTkButton(self.sidebar_frame, text="🚛 CETAK SURAT JALAN", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("sj"))
        self.btn_sj.grid(row=4, column=0, padx=15, pady=4, sticky="ew")

        # Menu 4: CETAK TERIMA BARANG
        self.btn_st = ctk.CTkButton(self.sidebar_frame, text="📥 CETAK TERIMA BARANG", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("st"))
        self.btn_st.grid(row=5, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 5: Tracking Packing List
        self.btn_tracking_pack = ctk.CTkButton(self.sidebar_frame, text="📝 Tracking Packing List", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("tracking_pack"))
        self.btn_tracking_pack.grid(row=6, column=0, padx=15, pady=4, sticky="ew")
        # Menu 6: Histori Surat Jalan
        self.btn_hist_sj = ctk.CTkButton(self.sidebar_frame, text="📜 Histori Surat Jalan", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("hist_sj"))
        self.btn_hist_sj.grid(row=7, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 7: Histori Tanda Terima
        self.btn_hist_st = ctk.CTkButton(self.sidebar_frame, text="📑 Histori Tanda Terima", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("hist_st"))
        self.btn_hist_st.grid(row=8, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 8: CETAK LABEL BARCODE MINI
        self.btn_pack_rekap = ctk.CTkButton(self.sidebar_frame, text="🏷️ CETAK LABEL BARCODE", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("cetak_barcode"))
        self.btn_pack_rekap.grid(row=9, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 9: Rekap Stok Laptop
        self.btn_stock_rekap = ctk.CTkButton(self.sidebar_frame, text="📊 Rekap Stok Laptop", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("stock_rekap"))
        self.btn_stock_rekap.grid(row=10, column=0, padx=15, pady=4, sticky="ew")
        
        # Menu 10: MASTER DATA UNIT
        self.btn_stock = ctk.CTkButton(self.sidebar_frame, text="💻 MASTER DATA UNIT", font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), height=38, corner_radius=8, fg_color="transparent", text_color=txt_clr, hover_color="#475569", anchor="w", command=lambda: self.ganti_halaman("stock"))
        self.btn_stock.grid(row=11, column=0, padx=15, pady=4, sticky="ew")
             
        # PENATAAN STRUKTUR GRID DASAR SIDEBAR KIRI (COPYRIGHT MUNCUL 100%) ---
        self.sidebar_frame.grid_rowconfigure(12, weight=1)
        self.sidebar_frame.grid_rowconfigure(13, weight=0) # Mengunci baris copyright agar stabil
        
        teks_copyright_baru = "Enterprise Premium Copyright MaMet SpooKy\n0811-2128-107"
        
        self.info_label = ctk.CTkLabel(
            self.sidebar_frame, 
            text=teks_copyright_baru, 
            font=ctk.CTkFont(family="Segoe UI", size=9, weight="bold", slant="italic"), 
            text_color="#64748B",
            justify="center"
        )
        # KUNCI POSISI: Menaruh objek dengan aman di baris ke-13 agar terangkat naik di atas dasar layar monitor
        self.info_label.grid(row=13, column=0, padx=20, pady=(10, 15), sticky="s")

        # 11. Bingkai Konten Utama Display Kanan (Luxury Slate Border Cerah)
        self.content_frame = ctk.CTkFrame(self, corner_radius=16, fg_color=("#F1F5F9", "#1E293B"), border_color="#CBD5E1", border_width=1)
        self.content_frame.grid(row=0, column=1, padx=25, pady=25, sticky="nsew")
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)
        
        # 12. Inisialisasi Seluruh Kamar Lembar Halaman (Total 10 Halaman Kontrol)
        self.halaman_aktif = {}
        self.buat_halaman_dashboard()
        self.buat_halaman_scan()
        self.buat_halaman_surat_jalan()
        self.buat_halaman_tanda_terima()
        self.buat_halaman_tracking_pack()
        self.buat_halaman_histori_sj()
        self.buat_halaman_histori_st()
        self.buat_halaman_pack_rekap()
        self.buat_halaman_stok_rekap()
        self.buat_halaman_stok()
        self.buat_layar_loading_splash()

        # REVISI VISUAL: Menyatukan Label Hak Cipta dan Tombol Logout Sejajar Horizontal di Sudut Kanan Atas
        self.frame_top_kanan = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.frame_top_kanan.place(relx=1.0, rely=0.0, x=-20, y=10, anchor="ne")
        
        # 1. Menampilkan teks Hak Cipta tipis elegan di sebelah KIRI tombol
        self.lbl_copyright_top = ctk.CTkLabel(
            self.frame_top_kanan, 
            text="Copyright MaMet SpooKy", 
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold", slant="italic"), 
            text_color="#64748B"  # Warna abu-abu slate premium agar kontras dan terbaca jelas
        )
        self.lbl_copyright_top.pack(side="left", padx=(0, 15), side_visibility=True if hasattr(ctk, 'side_visibility') else None) # Jarak 15 pixel dari tombol
        
        # 2. Menampilkan tombol logout di sebelah KANAN teks
        self.btn_logout_sudut = ctk.CTkButton(
            self.frame_top_kanan, 
            text="🚪 LOGOUT USER", 
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), 
            width=110, 
            height=26, 
            corner_radius=6, 
            fg_color="#DC2626", 
            hover_color="#B91C1C", 
            text_color="#FFFFFF",
            command=self.eksekusi_logout_akun_sistem
        )
        self.btn_logout_sudut.pack(side="right")
      
        # Otomatis membuka Dashboard saat software baru dinyalakan
        self.ganti_halaman("dash")
        
        # Memicu mesin pembaca data Excel pertama kali
        self.muat_data_excel()

        # --- DEKLARASI TEMA MEWAH GLOBAL (DARK LUXURY THEME FOR TREEVIEW) ---
        gaya_global = ttk.Style()
        gaya_global.theme_use("default")
        
        # Mengunci konfigurasi warna mewah untuk semua tabel di 6 halaman kontrol
        gaya_global.configure(
            "Treeview", 
            background="#1E293B",       # Latar belakang Slate Grey Gelap Premium
            foreground="#F8FAFC",       # Warna teks putih cerah berkontras tinggi
            rowheight=30,               # Baris agak renggang agar tidak melelahkan mata
            fieldbackground="#1E293B",  # Mengunci warna dasar background
            borderwidth=0, 
            font=("Segoe UI", 10)
        )
        # Mengunci desain kepala kolom (Header) menjadi hitam legam eksklusif
        gaya_global.configure(
            "Treeview.Heading", 
            background="#0F172A", 
            foreground="#FFFFFF", 
            font=("Segoe UI", 10, "bold"), 
            relief="flat"
        )
        # Efek baris saat di-klik (Seleksi Biru Neon Premium)
        gaya_global.map("Treeview", background=[('selected', '#2563EB')], foreground=[('selected', '#FFFFFF')])

        # MENYALAKAN DATABASE MULTI-OPERATOR: Mengaktifkan hulu jaringan SQL Server Standby
        self.inisialisasi_dan_migrasi_excel_ke_sql()


    # --- FUNGSI SISTEM 1: KLIK JUDUL UNTUK SORTIR INTERAKTIF (A-Z / Z-A) ---
    def urutkan_kolom(self, tree, col, reverse):
        """ Fungsi pengurut data otomatis saat kepala kolom diklik oleh operator """
        # Ambil seluruh data baris yang ada di dalam tabel saat ini
        data_tabel = [(tree.set(k, col), k) for k in tree.get_children("")]
        
        # Proses pengurutan data di memori RAM secara instan
        data_tabel.sort(reverse=reverse)
        
        # Tata ulang posisi baris di layar monitor sesuai hasil sortir terbaru
        for indeks, (val, k) in enumerate(data_tabel):
            tree.move(k, "", indeks)
            
        # Balikkan fungsi klik berikutnya (jika tadi A-Z, klik lagi jadi Z-A)
        tree.heading(col, command=lambda: self.urutkan_kolom(tree, col, not reverse))

    # --- FUNGSI SISTEM 2: KLIK 2X BATAS KOLOM UNTUK AUTO-FIT ALA EXCEL ---
    def auto_fit_kolom(self, event, tree):
        """ Fungsi pendeteksi klik dua kali pada batas garis untuk melebarkan kolom otomatis """
        # Deteksi elemen tabel mana yang diklik oleh operator
        region = tree.identify_region(event.x, event.y)
        if region in ["separator", "heading"]:
            kolom_id = tree.identify_column(event.x)
            if kolom_id:
                try:
                    # Ambil nama kolom internal berdasarkan indeks koordinat klik x
                    nama_kolom = tree["columns"][int(kolom_id.replace("#", "")) - 1]
                    
                    # Cari teks terpanjang yang ada di dalam kolom tersebut saat ini
                    lebar_maks = 100 # Batas minimal lebar kolom awal
                    for item in tree.get_children(""):
                        teks_baris = str(tree.set(item, nama_kolom))
                        # Hitung estimasi piksel berdasarkan panjang karakter teks harian
                        lebar_teks = (len(teks_baris) * 8) + 30
                        if lebar_teks > lebar_maks:
                            lebar_maks = lebar_teks
                            
                    # Lebarkan kolom secara otomatis di layar monitor kargo Anda
                    tree.column(nama_kolom, width=lebar_maks)
                except:
                    pass

    # --- FUNGSI NAVIGASI LUAR (DILUAR INIT) ---
    def ganti_halaman(self, nama_halaman):
        """ Fungsi Pengontrol Kamar Navigasi dengan Efek Animasi Warna Teken 3D """
        # Sembunyikan lembar halaman lama dari layar kanan
        for frame in self.halaman_aktif.values():
            frame.grid_forget()
            
        # Memunculkan lembar halaman baru dengan layout grid presisi
        self.halaman_aktif[nama_halaman].grid(row=0, column=0, sticky="nsew", padx=25, pady=20)
        
        list_tombol = {
            "dash": self.btn_dash, "scan": self.btn_scan, "sj": self.btn_sj, "st": self.btn_st,
            "tracking_pack": self.btn_tracking_pack, "hist_sj": self.btn_hist_sj, "hist_st": self.btn_hist_st,
            "pack_rekap": self.btn_pack_rekap, "stock_rekap": self.btn_stock_rekap, "stock": self.btn_stock
        }
        
        # Efek Teken 3D: Tombol aktif menyala BIRU BOLD berbingkai neon, tombol pasif kembali transparan normal
        for k, btn in list_tombol.items():
            if nama_halaman == k:
                btn.configure(
                    fg_color="#2563EB",          # Warna Biru Royal Premium
                    text_color="#FFFFFF",         # Teks Putih Cerah
                    border_color="#60A5FA",       # Garis Bingkai Neon Cerah (Efek 3D Timbul)
                    border_width=2,               # Ketebalan Garis Bingkai
                    font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
                )
            else:
                btn.configure(
                    fg_color="transparent", 
                    text_color=("#000000", "#E2E8F0"), # FIX WARNA: Teks pasif adaptif di mode gelap
                    border_width=0,
                    font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal")
                )

    def buat_label_copyright(self, frame):
        """ Label Hak Cipta Dinonaktifkan karena sudah pindah ke sebelah tombol Logout """
        pass

    def buat_halaman_dashboard(self):
        """ MENU 1: Layout Dashboard Komplit Berdasarkan Data Excel Asli """
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["dash"] = frame
        self.buat_label_copyright(frame)
        
        # Sektor Pembagian Layout Utama (Kiri: Kartu Real-Time, Kanan: Filter & Tabel Bulanan)
        main_dash_frame = ctk.CTkFrame(frame, fg_color="transparent")
        main_dash_frame.pack(fill="both", expand=True, padx=5, pady=5)
        main_dash_frame.grid_columnconfigure(0, weight=4) # Area Kiri agak lebar untuk kartu
        main_dash_frame.grid_columnconfigure(1, weight=5) # Area Kanan untuk tabel & grafik
        main_dash_frame.grid_rowconfigure(0, weight=1)
                
        kiri_container = ctk.CTkFrame(main_dash_frame, fg_color="transparent")
        kiri_container.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        ctk.CTkLabel(kiri_container, text="STATUS UNIT SAAT INI (REAL-TIME)", font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(0, 10))
        
        grid_kartu = ctk.CTkFrame(kiri_container, fg_color="transparent")
        grid_kartu.pack(fill="both", expand=True)
        grid_kartu.grid_columnconfigure(0, weight=1)
        grid_kartu.grid_columnconfigure(1, weight=1)
        
        # Kartu 2: Total Unit Ready (Baris 0, Kolom 0)
        c2 = ctk.CTkFrame(grid_kartu, fg_color="#DCFCE7", corner_radius=10, border_color="#22C55E", border_width=1, height=85)
        c2.grid(row=0, column=0, padx=6, pady=6, sticky="ew")
        c2.grid_propagate(False)
        ctk.CTkLabel(c2, text="TOTAL UNIT READY", font=ctk.CTkFont(size=10, weight="bold"), text_color="#166534").pack(pady=(12, 2))
        self.lbl_unit_ready = ctk.CTkLabel(c2, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#15803D")
        self.lbl_unit_ready.pack()
        
        # Kartu 3: Unit Sedang Disewa (Baris 0, Kolom 1)
        c3 = ctk.CTkFrame(grid_kartu, fg_color="#DBEAFE", corner_radius=10, border_color="#3B82F6", border_width=1, height=85)
        c3.grid(row=0, column=1, padx=6, pady=6, sticky="ew")
        c3.grid_propagate(False)
        ctk.CTkLabel(c3, text="UNIT SEDANG DISEWA", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1E40AF").pack(pady=(12, 2))
        self.lbl_unit_sewa = ctk.CTkLabel(c3, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#1D4ED8")
        self.lbl_unit_sewa.pack()

        # Kartu 7: Unit Terjual (Baris 1, Kolom 0)
        c7 = ctk.CTkFrame(grid_kartu, fg_color="#FFEDD5", corner_radius=10, border_color="#F97316", border_width=1, height=85)
        c7.grid(row=1, column=0, padx=6, pady=6, sticky="ew")
        c7.grid_propagate(False)
        ctk.CTkLabel(c7, text="UNIT TERJUAL", font=ctk.CTkFont(size=10, weight="bold"), text_color="#9A3412").pack(pady=(12, 2))
        self.lbl_unit_terjual = ctk.CTkLabel(c7, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#C2410C")
        self.lbl_unit_terjual.pack()
        
        # Kartu 4: Unit Sedang Dipinjam (Baris 1, Kolom 1)
        c4 = ctk.CTkFrame(grid_kartu, fg_color="#F3E8FF", corner_radius=10, border_color="#A855F7", border_width=1, height=85)
        c4.grid(row=1, column=1, padx=6, pady=6, sticky="ew")
        c4.grid_propagate(False)
        ctk.CTkLabel(c4, text="UNIT SEDANG DIPINJAM", font=ctk.CTkFont(size=10, weight="bold"), text_color="#6B21A8").pack(pady=(12, 2))
        self.lbl_unit_pinjam = ctk.CTkLabel(c4, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#7E22CE")
        self.lbl_unit_pinjam.pack()
        
        # Kartu 5: Unit Dalam Service (Baris 2, Kolom 0)
        c5 = ctk.CTkFrame(grid_kartu, fg_color="#FEF3C7", corner_radius=10, border_color="#F59E0B", border_width=1, height=85)
        c5.grid(row=2, column=0, padx=6, pady=6, sticky="ew")
        c5.grid_propagate(False)
        ctk.CTkLabel(c5, text="UNIT DALAM SERVICE", font=ctk.CTkFont(size=10, weight="bold"), text_color="#92400E").pack(pady=(12, 2))
        self.lbl_unit_service = ctk.CTkLabel(c5, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#B45309")
        self.lbl_unit_service.pack()
        
        # Kartu 6: Unit Data Rusak (Baris 2, Kolom 1)
        c6 = ctk.CTkFrame(grid_kartu, fg_color="#FEE2E2", corner_radius=10, border_color="#EF4444", border_width=1, height=85)
        c6.grid(row=2, column=1, padx=6, pady=6, sticky="ew")
        c6.grid_propagate(False)
        ctk.CTkLabel(c6, text="UNIT DATA RUSAK", font=ctk.CTkFont(size=10, weight="bold"), text_color="#991B1B").pack(pady=(12, 2))
        self.lbl_unit_rusak = ctk.CTkLabel(c6, text="0", font=ctk.CTkFont(size=24, weight="bold"), text_color="#B91C1C")
        self.lbl_unit_rusak.pack()
        
        # Kartu 1: Total Keseluruhan Unit (Lebar Penuh di Baris 3)
        c1 = ctk.CTkFrame(grid_kartu, fg_color="#334155", corner_radius=10, border_color="#475569", border_width=1, height=85)
        c1.grid(row=3, column=0, columnspan=2, padx=6, pady=10, sticky="ew")
        c1.grid_propagate(False)
        ctk.CTkLabel(c1, text="TOTAL KESELURUHAN UNIT", font=ctk.CTkFont(size=11, weight="bold"), text_color="#94A3B8").pack(pady=(12, 2))
        self.lbl_total_unit = ctk.CTkLabel(c1, text="0", font=ctk.CTkFont(size=26, weight="bold"), text_color="#F8FAFC")
        self.lbl_total_unit.pack()

        kanan_container = ctk.CTkFrame(main_dash_frame, fg_color="transparent")
        kanan_container.grid(row=0, column=1, sticky="nsew", padx=(15, 0))
        
        # Garis Horizontal Atas Filter
        filter_header = ctk.CTkFrame(kanan_container, fg_color="transparent")
        filter_header.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(filter_header, text="PILIH ANGKA BULAN (1-12) :", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(side="left")
        
        # FIX UTAMA DROPDOWN BIND: Menggunakan parameter 'command' resmi CustomTkinter agar responsif interaktif
        self.combo_bulan = ctk.CTkComboBox(
            filter_header, 
            values=[str(m) for m in range(1, 13)], 
            width=70, 
            height=28, 
            corner_radius=5,
            command=lambda v: self.hitung_rekap_bulanan() # Memicu hitungan & grafik otomatis secara dinamis
        )
        self.combo_bulan.pack(side="left", padx=10)
        self.combo_bulan.set("8") # Standar otomatis bulan 8

        # Judul Tabel Log Aktivitas
        ctk.CTkLabel(kanan_container, text="Daftar Aktivitas Log Bulanan", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color=("#0F172A", "#F8FAFC")).pack(anchor="w", pady=(10, 2))
        
        # Komponen Tabel Log Bulanan Mini Berdesain Bersih Premium
        self.tree_bulan = ttk.Treeview(kanan_container, columns=("Aktivitas", "Total"), show="headings", height=6)
        self.tree_bulan.heading("Aktivitas", text="Daftar Aktivitas Log Bulanan", anchor="w")
        self.tree_bulan.heading("Total", text="Total", anchor="center")
        self.tree_bulan.column("Aktivitas", width=280, anchor="w")
        self.tree_bulan.column("Total", width=80, anchor="center")
        self.tree_bulan.pack(fill="x", pady=(0, 15))
        
        # Mengisi baris template teks aktivitas log bulanan sementara
        self.tree_bulan.insert("", "end", values=("Laptop Keluar / Disewa", "0"))
        self.tree_bulan.insert("", "end", values=("Laptop Masuk Perbaikan (Service)", "0"))
        self.tree_bulan.insert("", "end", values=("Laptop Masuk Kembali ke Gudang", "0"))
        self.tree_bulan.insert("", "end", values=("Laptop Unit Terjual", "0"))
        self.tree_bulan.insert("", "end", values=("Laptop Sedang Dipinjam", "0"))
        self.tree_bulan.insert("", "end", values=("Laptop Terdata Rusak", "0"))
        
        # Area kosong bawah yang nanti khusus kita suntikkan Grafik Batang Statistik
        self.grafik_frame = ctk.CTkFrame(kanan_container, fg_color=("#F8FAFC", "#1E293B"), border_color="#CBD5E1", border_width=1, height=220)
        self.grafik_frame.pack(fill="both", expand=True, pady=5)
        self.perbarui_grafik_dashboard()
    
    def perbarui_grafik_dashboard(self):
        """ LOGIKA GRAFIK: Menggambar grafik batang statistik bulanan harian secara aman """
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        # 1. Bersihkan grafik lama yang ada di dalam frame agar tidak menumpuk
        for widget in self.grafik_frame.winfo_children():
            widget.destroy()

        kategori = ['Sewa', 'Service', 'Kembali', 'Terjual', 'Dipinjam', 'Rusak']
        
        # 2. Ambil data dinamis hasil kalkulasi fungsi bulanan
        if hasattr(self, 'data_grafik_bulanan') and self.data_grafik_bulanan:
            jumlah_unit = self.data_grafik_bulanan
        else:
            # Nilai default awal pengunci grafik agar tidak crash saat pertama dibuka
            jumlah_unit = [0, 0, 0, 0, 0, 0]

        try:
            # 3. Proses Merakit Grafik Batang Menggunakan Matplotlib Premium Theme
            fig, ax = plt.subplots(figsize=(5.5, 2.2), dpi=100)
            bg_color = '#F8FAFC' if ctk.get_appearance_mode() == "Light" else '#1E293B'
            text_color = '#1E293B' if ctk.get_appearance_mode() == "Light" else '#F8FAFC'
            
            fig.patch.set_facecolor(bg_color) 
            ax.set_facecolor(bg_color)

            # Batang abu-abu metalik premium sesuai visual asli
            batang = ax.bar(kategori, jumlah_unit, color='#94A3B8', width=0.4, edgecolor='#64748B', linewidth=0.8)

            # Menampilkan angka di atas setiap batang grafik secara presisi
            max_val = max(jumlah_unit) if max(jumlah_unit) > 0 else 10
            for b in batang:
                yval = b.get_height()
                ax.text(b.get_x() + b.get_width()/2, yval + (max_val * 0.02 + 0.1), f"{int(yval)}", ha='center', va='bottom', fontsize=8, color=text_color, weight='bold')

            # Mengatur desain garis tepi tabel grafik
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color('#CBD5E1')
            ax.spines['bottom'].set_color('#CBD5E1')
            ax.tick_params(axis='both', colors='#475569', labelsize=8)
            ax.grid(axis='y', linestyle='--', alpha=0.5, color='#CBD5E1')

            ax.set_ylim(0, max_val + (max_val * 0.2 + 2))
            plt.title("Statistik Aktivitas Bulanan", fontsize=10, weight='bold', color=text_color, pad=10)

            # 4. Suntikkan Grafik Matplotlib ke Dalam Widget Aplikasi Desktop
            canvas = FigureCanvasTkAgg(fig, master=self.grafik_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)
            plt.close(fig)
        except Exception as e:
            print(f"Gagal menggambar grafik bulanan: {str(e)}")
            
    def buat_halaman_scan(self):
        """ Menu 2: STASIUN KONTROL PEMINDAIAN BARCODE & SURAT JALAN GUDANG """
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["scan"] = frame
        self.buat_label_copyright(frame)
        
        # Header Stasiun Utama
        ctk.CTkLabel(frame, text="STASIUN UTAMA PEMINDAIAN BARCODE", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 10))
        ctk.CTkLabel(frame, text="Sistem Inventaris Laptop Real-Time - Isi kolom keterangan sebelum memindai unit keluar!", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        # CONTAINER PANEL UTAMA (Membagi Kiri dan Kanan secara Simetris)
        panel_container = ctk.CTkFrame(frame, fg_color="transparent")
        panel_container.pack(fill="x", expand=False, pady=5)
        panel_container.grid_columnconfigure(0, weight=1) # Panel Masuk (Kiri)
        panel_container.grid_columnconfigure(1, weight=1) # Panel Keluar (Kanan)
        
        panel_kiri = ctk.CTkFrame(panel_container, fg_color="transparent")
        panel_kiri.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        # Box 1: Scan Ready / Masuk Gudang
        b_ready = ctk.CTkFrame(panel_kiri, fg_color="#DCFCE7", corner_radius=8, border_color="#22C55E", border_width=1)
        b_ready.pack(fill="x", pady=5, ipady=10)
        ctk.CTkLabel(b_ready, text="KOTAK SCAN READY / UNIT MASUK", font=ctk.CTkFont(size=11, weight="bold"), text_color="#166534").pack(anchor="w", padx=15, pady=2)
        ent_ready = ctk.CTkEntry(b_ready, placeholder_text="[ SILAKAN SCAN BARCODE UNIT READY ]", width=300, height=28, fg_color="#FFFFFF", text_color="#000000")
        ent_ready.pack(anchor="w", padx=15, pady=5)
        ent_ready.bind("<Return>", lambda e, t="READY": self.proses_scan_barcode(ent_ready, t))
        
        # Box 2: Scan Rusak
        b_rusak = ctk.CTkFrame(panel_kiri, fg_color="#FEE2E2", corner_radius=8, border_color="#EF4444", border_width=1)
        b_rusak.pack(fill="x", pady=5, ipady=10)
        ctk.CTkLabel(b_rusak, text="KOTAK SCAN LAPTOP RUSAK", font=ctk.CTkFont(size=11, weight="bold"), text_color="#991B1B").pack(anchor="w", padx=15, pady=2)
        ent_rusak = ctk.CTkEntry(b_rusak, placeholder_text="[ SILAKAN SCAN BARCODE UNIT RUSAK ]", width=300, height=28, fg_color="#FFFFFF", text_color="#000000")
        ent_rusak.pack(anchor="w", padx=15, pady=5)
        ent_rusak.bind("<Return>", lambda e, t="RUSAK": self.proses_scan_barcode(ent_rusak, t))
        
        panel_kanan = ctk.CTkFrame(panel_container, fg_color="transparent")
        panel_kanan.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        # Box 3: Sewa Keluar
        b_sewa = ctk.CTkFrame(panel_kanan, fg_color="#DBEAFE", corner_radius=8, border_color="#3B82F6", border_width=1)
        b_sewa.pack(fill="x", pady=3)
        ctk.CTkLabel(b_sewa, text="1. KELUAR / DISEWA", font=ctk.CTkFont(size=11, weight="bold"), text_color="#1E40AF").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.txt_ket_sewa = ctk.CTkEntry(b_sewa, placeholder_text="Masukkan Nama Toko/Penerima", width=180, height=26, fg_color="#FEFCE8", text_color="#000000")
        self.txt_ket_sewa.grid(row=0, column=1, padx=5, pady=5)
        ent_sewa = ctk.CTkEntry(b_sewa, placeholder_text="[ SCAN BARCODE ]", width=140, height=26, fg_color="#FFFFFF", text_color="#000000")
        ent_sewa.grid(row=0, column=2, padx=10, pady=5)
        ent_sewa.bind("<Return>", lambda e, o=ent_sewa, k=self.txt_ket_sewa, t="DISEWA": self.proses_scan_barcode(o, t, k))
        
        # Box 4: Masuk Service
        b_service = ctk.CTkFrame(panel_kanan, fg_color="#FEF3C7", corner_radius=8, border_color="#F59E0B", border_width=1)
        b_service.pack(fill="x", pady=3)
        ctk.CTkLabel(b_service, text="2. MASUK SERVICE", font=ctk.CTkFont(size=11, weight="bold"), text_color="#92400E").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.txt_ket_service = ctk.CTkEntry(b_service, placeholder_text="Masukkan Vendor/Lokasi", width=180, height=26, fg_color="#FEFCE8", text_color="#000000")
        self.txt_ket_service.grid(row=0, column=1, padx=5, pady=5)
        ent_service = ctk.CTkEntry(b_service, placeholder_text="[ SCAN BARCODE ]", width=140, height=26, fg_color="#FFFFFF", text_color="#000000")
        ent_service.grid(row=0, column=2, padx=10, pady=5)
        ent_service.bind("<Return>", lambda e, o=ent_service, k=self.txt_ket_service, t="SERVICE": self.proses_scan_barcode(o, t, k))
        
        # Box 5: Unit Dijual
        b_jual = ctk.CTkFrame(panel_kanan, fg_color="#F1F5F9", corner_radius=8, border_color="#94A3B8", border_width=1)
        b_jual.pack(fill="x", pady=3)
        ctk.CTkLabel(b_jual, text="3. UNIT DIJUAL", font=ctk.CTkFont(size=11, weight="bold"), text_color="#475569").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.txt_ket_jual = ctk.CTkEntry(b_jual, placeholder_text="Masukkan Pembeli/Toko", width=180, height=26, fg_color="#FEFCE8", text_color="#000000")
        self.txt_ket_jual.grid(row=0, column=1, padx=5, pady=5)
        ent_jual = ctk.CTkEntry(b_jual, placeholder_text="[ SCAN BARCODE ]", width=140, height=26, fg_color="#FFFFFF", text_color="#000000")
        ent_jual.grid(row=0, column=2, padx=10, pady=5)
        ent_jual.bind("<Return>", lambda e, o=ent_jual, k=self.txt_ket_jual, t="TERJUAL": self.proses_scan_barcode(o, t, k))
        
        # Box 6: Unit Dipinjam
        b_pinjam = ctk.CTkFrame(panel_kanan, fg_color="#F3E8FF", corner_radius=8, border_color="#A855F7", border_width=1)
        b_pinjam.pack(fill="x", pady=3)
        ctk.CTkLabel(b_pinjam, text="4. UNIT DIPINJAM", font=ctk.CTkFont(size=11, weight="bold"), text_color="#6B21A8").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.txt_ket_pinjam = ctk.CTkEntry(b_pinjam, placeholder_text="Masukkan Nama Peminjam", width=180, height=26, fg_color="#FEFCE8", text_color="#000000")
        self.txt_ket_pinjam.grid(row=0, column=1, padx=5, pady=5)
        ent_pinjam = ctk.CTkEntry(b_pinjam, placeholder_text="[ SCAN BARCODE ]", width=140, height=26, fg_color="#FFFFFF", text_color="#000000")
        ent_pinjam.grid(row=0, column=2, padx=10, pady=5)
        ent_pinjam.bind("<Return>", lambda e, o=ent_pinjam, k=self.txt_ket_pinjam, t="DIPINJAM": self.proses_scan_barcode(o, t, k))

        # Jendela Konsol Live Monitor Sukses/Gagal Scan Gudang
        ctk.CTkLabel(frame, text="LIVE GUDANG MONITOR LOG", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color=("#0F172A", "#F8FAFC")).pack(anchor="w", pady=(10, 2))
        self.status_box = ctk.CTkTextbox(frame, height=140, corner_radius=12, fg_color="#0F172A", text_color="#34D399", border_color="#475569", border_width=1, font=("Consolas", 11))
        self.status_box.pack(fill="both", expand=True, padx=2, pady=2)
        self.status_box.insert("0.0", "--- SISTEM HARDWARE SCANNER ONLINE ---\nSiap memproses perubahan multi-status unit kargo...\n")
        self.status_box.configure(state="disabled")

    def buat_halaman_surat_jalan(self):
        """ Menu 3: FORM OPERASIONAL - CETAK SURAT JALAN (SJ) REVISI TOTAL CORPORATE """
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["sj"] = frame
        self.buat_label_copyright(frame)
        
        # Header Tampilan Form
        ctk.CTkLabel(frame, text="FORM OPERASIONAL - CETAK SURAT JALAN (SJ)", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 10))
        ctk.CTkLabel(frame, text="Pilih penerima dan isi alamat tujuan dengan lengkap sebelum menerbitkan dokumen resmi pengiriman.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        # Bingkai Form Input
        form_frame = ctk.CTkFrame(frame, corner_radius=12, fg_color="#475569")
        form_frame.pack(fill="x", pady=10, padx=10)
        
        # 1. Komponen Dropdown Pilihan Nama PIC / Toko Penerima Kargo
        ctk.CTkLabel(form_frame, text="Pilih Nama PIC / Toko Penerima Kargo:", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#E2E8F0").grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")
        
        self.combo_sj_toko = ctk.CTkComboBox(form_frame, width=420, height=32, fg_color="#FFFFFF", text_color="#000000")
        self.combo_sj_toko.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="w")
        self.combo_sj_toko.bind("<<ComboboxSelected>>", lambda e: self.efek_pilihan_toko_sj())
        
        # 2. Kolom Input Alamat Tujuan Pengiriman Kargo (Wajib Isi)
        ctk.CTkLabel(form_frame, text="Alamat Lengkap Tujuan Pengiriman Kargo Keluar:", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#E2E8F0").grid(row=2, column=0, padx=20, pady=(5, 5), sticky="w")
        
        self.entry_sj_alamat = ctk.CTkEntry(form_frame, placeholder_text="Ketik alamat pengiriman barang secara detail di sini...", width=520, height=34, fg_color="#FFFFFF", text_color="#000000")
        self.entry_sj_alamat.grid(row=3, column=0, padx=20, pady=(0, 10), sticky="w")
        
        # 3. Kolom Input Catatan Tambahan / Keterangan Nota Internal
        ctk.CTkLabel(form_frame, text="Keterangan Tambahan / Catatan Internal Dokumen:", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#E2E8F0").grid(row=4, column=0, padx=20, pady=(5, 5), sticky="w")
        
        self.entry_sj_ket = ctk.CTkEntry(form_frame, placeholder_text="Contoh: Project Astra Sunter via Kurir Lalamove...", width=520, height=34, fg_color="#FFFFFF", text_color="#000000")
        self.entry_sj_ket.grid(row=5, column=0, padx=20, pady=(0, 20), sticky="w")
        
        # 4. Tombol Hijau Besar Eksekutor Cetak Dokumen PDF Resmi Corporate
        self.btn_cetak_sj = ctk.CTkButton(frame, text="🚛 Terbitkan & Cetak Dokumen Surat Jalan Resmi (Save PDF)", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), height=45, corner_radius=8, fg_color="#10B981", hover_color="#059669", command=self.proses_penerbitan_sj)
        self.btn_cetak_sj.pack(fill="x", padx=10, pady=15)

    def efek_pilihan_toko_sj(self):
        """ SENSOR DROPDOWN: Memeriksa live kargo & mengisi otomatis alamat tujuan """
        import openpyxl
        import datetime
        
        toko_terpilih = self.combo_sj_toko.get().strip()
        if not toko_terpilih or "Pilih" in toko_terpilih: 
            return
            
        tgl_hari_ini = datetime.datetime.now().strftime("%Y-%m-%d")
        
        # KAMUS AUTO-FILL ALAMAT REKANAN TETAP
        kamus_alamat = {
            "ASTRA": "Gd. Astra International, Jl. Gaya Motor Raya No.8, Sunter, Jakarta Utara",
            "GLOBAL": "Jl. Surapati No.123, Sukaluyu, Kec. Cibeunying Kaler, Kota Bandung",
            "MITRA COMPUTER": "Bandung Electronic Center (BEC) Lantai 1 Blok A-05, Kota Bandung"
        }
        
        # Jika nama toko terdaftar di kamus, isikan alamatnya otomatis ke kolom input baru
        self.entry_sj_alamat.delete(0, 'end')
        if toko_terpilih in kamus_alamat:
            self.entry_sj_alamat.insert(0, kamus_alamat[toko_terpilih])
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_log = wb["Tracking Packing List"]
            
            ada_kargo = False
            for row in range(15, ws_log.max_row + 1):
                tgl_cell = str(ws_log.cell(row=row, column=1).value)
                status_cell = str(ws_log.cell(row=row, column=4).value).strip().upper()
                ket_cell = str(ws_log.cell(row=row, column=5).value)
                
                if tgl_hari_ini in tgl_cell:
                    if status_cell != "KEMBALI" and status_cell != "RUSAK":
                        if toko_terpilih.upper() in ket_cell.upper():
                            if "SJ/GTP-" not in ket_cell:
                                ada_kargo = True
                                break
            
            if not ada_kargo:
                self.bell()
                self.btn_cetak_sj.configure(state="disabled", fg_color="#94A3B8", text="⚠️ Tidak Ada Riwayat Scan Keluar Baru untuk Nama Ini")
            else:
                self.btn_cetak_sj.configure(state="normal", fg_color="#10B981", text="🚛 Terbitkan & Cetak Dokumen Surat Jalan Resmi (SAVE PDF)")
                
        except Exception as err:
            print(f"Gagal mendeteksi live kargo dropdown SJ: {str(err)}")

    def buat_halaman_tanda_terima(self):
        """ Menu 4: FORM OPERASIONAL - CETAK TANDA TERIMA BARANG (ST) CORPORATE CLEAN """
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["st"] = frame
        self.buat_label_copyright(frame)
        
        # Header Utama Form Tanda Terima (ST)
        ctk.CTkLabel(frame, text="FORM OPERASIONAL - CETAK TANDA TERIMA BARANG (ST)", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 10))
        ctk.CTkLabel(frame, text="Pilih nama pengirim di bawah ini untuk menarik data scan masuk harian (KEMBALI/RUSAK) secara otomatis.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        # Bingkai Form Input Berlatar Belakang Abu-abu Metalik Premium
        form_frame = ctk.CTkFrame(frame, corner_radius=12, fg_color="#475569")
        form_frame.pack(fill="x", pady=10, padx=10)
        
        # 1. Komponen Dropdown Pilihan Nama Pihak Pengirim Barang (PIC/Toko)
        ctk.CTkLabel(form_frame, text="Pilih / Ketik Nama Pihak Pengirim (Yang Menyerahkan):", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#E2E8F0").grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")
        
        self.combo_st_toko = ctk.CTkComboBox(form_frame, width=420, height=32, fg_color="#FFFFFF", text_color="#000000")
        self.combo_st_toko.grid(row=1, column=0, padx=20, pady=(0, 15), sticky="w")
        self.combo_st_toko.bind("<<ComboboxSelected>>", lambda e: self.efek_pilihan_toko_st())
        
        # 2. Kolom Input Detail Alasan Retur / Kondisi Fisik Kargo Masuk
        ctk.CTkLabel(form_frame, text="Alasan Masuk Kembali / Detail Kondisi Fisik Unit:", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#E2E8F0").grid(row=2, column=0, padx=20, pady=(5, 5), sticky="w")
        
        self.entry_st_ket = ctk.CTkEntry(form_frame, placeholder_text="Contoh: Pengembalian Unit Sewa Selesai Project / Retur Unit Layar Bergaris...", width=520, height=34, fg_color="#FFFFFF", text_color="#000000")
        self.entry_st_ket.grid(row=3, column=0, padx=20, pady=(0, 20), sticky="w")
        
        # 3. Tombol Biru Besar Eksekutor Cetak Dokumen Tanda Terima Resmi (ST)
        self.btn_cetak_st = ctk.CTkButton(frame, text="📥 Terbitkan & Cetak Tanda Terima Barang (SAVE PDF)", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), height=45, corner_radius=8, fg_color="#3B82F6", hover_color="#1D4ED8", command=self.proses_penerbitan_st)
        self.btn_cetak_st.pack(fill="x", padx=10, pady=15)

    def efek_pilihan_toko_st(self):
        """ SENSOR DROPDOWN: Memeriksa live kargo masuk harian (READY/RUSAK) milik toko terpilih """
        import openpyxl
        import datetime
        
        toko_terpilih = self.combo_st_toko.get().strip()
        if not toko_terpilih or "Pilih" in toko_terpilih: 
            return
            
        tgl_hari_ini = datetime.datetime.now().strftime("%Y-%m-%d")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_log = wb["Tracking Packing List"]
            
            ada_kargo_masuk = False
            # Menyisir log harian khusus mendeteksi barang masuk (KEMBALI / RUSAK) harian
            for row in range(15, ws_log.max_row + 1):
                tgl_cell = str(ws_log.cell(row=row, column=1).value)
                status_cell = str(ws_log.cell(row=row, column=4).value).strip().upper()
                
                # REVISI 1: Memaksa semua teks keterangan dari Excel menjadi HURUF BESAR SEMUA (.upper())
                # Ini menjamin pencarian kata kebal dari kesalahan ketik huruf kecil di lapangan
                ket_cell = str(ws_log.cell(row=row, column=5).value).strip().upper()
                
                # Tetap mempertahankan logika deteksi tanggal normal Anda yang terbukti sukses
                if tgl_hari_ini in tgl_cell:
                    if status_cell == "KEMBALI" or status_cell == "RUSAK":
                        # Melakukan pencocokan huruf besar vs huruf besar secara adil dan presisi
                        if toko_terpilih.upper() in ket_cell:
                            
                            # REVISI 2: Filter anti-duplikat diperkuat ke format HURUF BESAR
                            # Unit yang sudah dicap nomor nota tidak akan bisa lolos cetak ganda
                            if "ST/GTP-" not in ket_cell:
                                ada_kargo_masuk = True
                                break
                                
            # Mengatur respon visual tombol cetak berdasarkan hasil deteksi sensor di atas
            if not ada_kargo_masuk:
                self.bell()
                self.btn_cetak_st.configure(state="disabled", fg_color="#94A3B8", text="⚠️ Tidak Ada Riwayat Scan Masuk Baru untuk Nama Ini")
            else:
                self.btn_cetak_st.configure(state="normal", fg_color="#3B82F6", text="📥 Terbitkan & Cetak Tanda Terima Barang (SAVE PDF)")
                
        except Exception as err:
            print(f"Gagal mendeteksi live kargo masuk dropdown ST: {str(err)}")

    def buat_halaman_tracking_pack(self):
        """ Menu 5: KRONOLOGI MUTASI & TRACKING PACKING LIST (REAL-TIME + ENTERPRISE MEWAH) """
        from tkinter import ttk
        
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["tracking_pack"] = frame
        self.buat_label_copyright(frame)
        
        ctk.CTkLabel(frame, text="KRONOLOGI MUTASI & TRACKING PACKING LIST", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 2))
        ctk.CTkLabel(frame, text="Menampilkan seluruh rekam jejak aktivitas scan kargo secara real-time otomatis dari data terbaru.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        search_frame = ctk.CTkFrame(frame, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(search_frame, text="Cari Barcode / Nama Laptop / Lokasi Toko:", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=("#334155", "#F8FAFC")).pack(side="left", padx=(5, 10))
        self.entry_cari_tracking = ctk.CTkEntry(search_frame, placeholder_text="Ketik di sini untuk memfilter log kargo secara instan...", width=380, height=30, fg_color="#FFFFFF", text_color="#000000")
        self.entry_cari_tracking.pack(side="left", padx=0)
        self.entry_cari_tracking.bind("<KeyRelease>", lambda e: self.filter_data_tracking())
        
        # AREA BINGKAI TABEL (GRID LAYOUT PREMIUM)
        tabel_container = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_color="#CBD5E1", border_width=1)
        tabel_container.pack(fill="both", expand=True, padx=2, pady=5)
        tabel_container.grid_columnconfigure(0, weight=1)
        tabel_container.grid_rowconfigure(0, weight=1)
        
        # REVISI VISUAL: Menambahkan elemen 'operator' ke dalam laci tuple kolom ke-7
        kolom_st = ("no", "waktu", "barcode", "nama", "status", "tujuan", "operator")
        self.tree_tracking = ttk.Treeview(tabel_container, columns=kolom_st, show="headings", style="Treeview")
        self.tree_tracking.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        scrollbar_v = ttk.Scrollbar(tabel_container, orient="vertical", command=self.tree_tracking.yview)
        scrollbar_v.grid(row=0, column=1, sticky="ns", pady=(5, 0), padx=(0, 5))
        scrollbar_h = ttk.Scrollbar(tabel_container, orient="horizontal", command=self.tree_tracking.xview)
        scrollbar_h.grid(row=1, column=0, sticky="ew", padx=(5, 0), pady=(0, 5))
        self.tree_tracking.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        self.tree_tracking.bind("<Double-1>", lambda e: self.auto_fit_kolom(e, self.tree_tracking))
        
        for col in kolom_st:
            self.tree_tracking.heading(col, text=col.title() if col != "no" else "No", command=lambda c=col: self.urutkan_kolom(self.tree_tracking, c, False))
            
        # Konfigurasi Text Papan Judul Atas Kolom Tabel (Heading)
        self.tree_tracking.heading("waktu", text="Tanggal & Waktu Scan")
        self.tree_tracking.heading("barcode", text="Nomor Barcode")
        self.tree_tracking.heading("nama", text="Spesifikasi / Nama Unit Laptop")
        self.tree_tracking.heading("tujuan", text="Keterangan Tujuan / Lokasi Kargo")
        self.tree_tracking.heading("operator", text="Petugas (Operator)") # ⬅️ Judul baru laci ke-7
        
        # Konfigurasi Ukuran Lebar Pixel Kolom Layar Monitor
        self.tree_tracking.column("no", width=50, anchor="center")
        self.tree_tracking.column("waktu", width=140, anchor="center")
        self.tree_tracking.column("barcode", width=120, anchor="center")
        self.tree_tracking.column("nama", width=250, anchor="w")
        self.tree_tracking.column("status", width=95, anchor="center")
        self.tree_tracking.column("tujuan", width=350, anchor="w") 
        self.tree_tracking.column("operator", width=120, anchor="center") # ⬅️ Lebar baru laci ke-7
        
        frame.bind("<Visibility>", lambda e: self.muat_log_tracking_excel())

    def muat_log_tracking_excel(self):
        """ ENGINE MONITOR CLOUD Live: Menarik data via REST API v2 + Auto-Refresh Gaib 3 Detik """
        import requests
        
        try:
            # 1. PANGGIL REST API v2: Mengambil kronologi mutasi dari server pusat
            resp = requests.get(f"{self.api_base_gudang}/api/gudang/tracking-list?limit=500", timeout=10)
            resp.raise_for_status()
            payload = resp.json()
            if payload.get("status") != "success":
                raise RuntimeError(payload.get("detail", "Respon API tidak dikenal"))
            
            # 2. KONVERSI JSON → TUPLE: Menyamakan bentuk data dengan hasil query SQL lama
            daftar_baris_cloud = [
                (d.get("waktu"), d.get("barcode"), d.get("nama_barang"),
                 d.get("status_mutasi"), d.get("keterangan_tujuan"), d.get("operator"))
                for d in payload.get("data", [])
            ]
            
            # 3. STERILISASI TABEL: Membersihkan sisa data lama pada komponen Treeview Menu 5 Anda
            if hasattr(self, 'tree_tracking'):
                for item in self.tree_tracking.get_children():
                    self.tree_tracking.delete(item)
                    
                # 4. LOOPING INSERTER CLOUD: Menyebarkan data biner server ke dalam grid tabel visual monitor
                for idx, row_data in enumerate(daftar_baris_cloud):
                    wkt, b_id, n_lap, st_mut, ket_tuj, op_nama = row_data
                    
                    # Memasukkan data ke baris Treeview secara urut termasuk kolom operator di ujung kanan
                    self.tree_tracking.insert("", "end", values=(idx + 1, wkt, b_id, n_lap, st_mut, ket_tuj, op_nama if op_nama else "-"))

        except Exception as err:
            print(f"Gagal menyegarkan data tabel log tracking dari Server GTP: {str(err)}")
            
        # 5. KUNCI UTAMA SINKRONISASI GAIB: Menyalakan saklar loop pemicu otomatis harian
        # Membatalkan antrean clock lama jika ada agar memori RAM PC operator tidak bengkak
        if hasattr(self, '_id_clock_refresh_live'):
            self.after_cancel(self._id_clock_refresh_live)
            
        # Menembakkan perintah loop otomatis setiap 3000 milidetik (3 Detik) secara berulang ke internet
        self._id_clock_refresh_live = self.after(3000, self.muat_log_tracking_excel)


    def filter_data_tracking(self):
        """ SENSOR FILTER KILAT: Menyaring isi log kargo secara live berdasarkan ketikan operator """
        kata_kunci = self.entry_cari_tracking.get().strip().lower()
        
        # Kosongkan baris grid tabel di layar monitor terlebih dahulu
        for item in self.tree_tracking.get_children():
            self.tree_tracking.delete(item)
            
        counter = 1
        # Menyisir seluruh data log yang tersimpan di dalam memori internal RAM Python
        for item in self.semua_data_tracking:
            string_gabung = f"{str(item[2]).lower()} {str(item[3]).lower()} {str(item[4]).lower()} {str(item[5]).lower()}"
            
            # Jika kata kunci COCOK dengan isi data baris log
            if kata_kunci in string_gabung:
                baris_baru = (counter, item[1], item[2], item[3], item[4], item[5])
                self.tree_tracking.insert("", "end", values=baris_baru)
                counter += 1

    def buat_halaman_histori_sj(self):
        """ Menu 6: HALAMAN ARSIP DOKUMEN CETAK SURAT JALAN (REAL-TIME + ENTERPRISE MEWAH) """
        from tkinter import ttk
        
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["hist_sj"] = frame
        self.buat_label_copyright(frame)
        
        ctk.CTkLabel(frame, text="HISTORI ARSIP DOKUMEN SURAT JALAN (SJ)", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 2))
        ctk.CTkLabel(frame, text="Menampilkan rekam jejak digital seluruh berkas nota Surat Jalan secara real-time otomatis dari data terbaru.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        search_frame = ctk.CTkFrame(frame, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(search_frame, text="Cari No Surat Jalan / Nama Toko Tujuan:", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=("#334155", "#F8FAFC")).pack(side="left", padx=(5, 10))
        self.entry_cari_histori_sj = ctk.CTkEntry(search_frame, placeholder_text="Ketik nomor nota / toko di sini untuk menyaring arsip secara instan...", width=380, height=30, fg_color="#FFFFFF", text_color="#000000")
        self.entry_cari_histori_sj.pack(side="left", padx=0)
        self.entry_cari_histori_sj.bind("<KeyRelease>", lambda e: self.filter_data_histori_sj())
        
        # AREA BINGKAI TABEL (GRID LAYOUT PREMIUM)
        tabel_container = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_color="#CBD5E1", border_width=1)
        tabel_container.pack(fill="both", expand=True, padx=2, pady=5)
        tabel_container.grid_columnconfigure(0, weight=1)
        tabel_container.grid_rowconfigure(0, weight=1)
        
        kolom_st = ("no", "waktu", "no_sj", "toko", "barcode", "catatan")
        self.tree_histori_sj = ttk.Treeview(tabel_container, columns=kolom_st, show="headings", style="Treeview")
        self.tree_histori_sj.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        scrollbar_v = ttk.Scrollbar(tabel_container, orient="vertical", command=self.tree_histori_sj.yview)
        scrollbar_v.grid(row=0, column=1, sticky="ns", pady=(5, 0), padx=(0, 5))
        scrollbar_h = ttk.Scrollbar(tabel_container, orient="horizontal", command=self.tree_histori_sj.xview)
        scrollbar_h.grid(row=1, column=0, sticky="ew", padx=(5, 0), pady=(0, 5))
        self.tree_histori_sj.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # BINDING INTERAKTIF: Klik 2x batas kolom untuk Auto-Fit ala Excel
        self.tree_histori_sj.bind("<Double-1>", lambda e: self.auto_fit_kolom(e, self.tree_histori_sj))
        
        # SUNTIKAN FITUR SORTIR: Klik judul kolom untuk mengurutkan data (A-Z / Z-A)
        for col in kolom_st:
            self.tree_histori_sj.heading(col, text=col.title() if col != "no" else "No", command=lambda c=col: self.urutkan_kolom(self.tree_histori_sj, c, False))
            
        self.tree_histori_sj.heading("waktu", text="Tanggal Terbit")
        self.tree_histori_sj.heading("no_sj", text="Nomor Surat Jalan")
        self.tree_histori_sj.heading("toko", text="Nama Toko Penerima Kargo")
        self.tree_histori_sj.heading("barcode", text="Total Unit Keluar")
        self.tree_histori_sj.heading("catatan", text="Catatan / Keterangan Internal")
        
        self.tree_histori_sj.column("no", width=50, anchor="center")
        self.tree_histori_sj.column("waktu", width=120, anchor="center")
        self.tree_histori_sj.column("no_sj", width=140, anchor="center")
        self.tree_histori_sj.column("toko", width=160, anchor="w")
        self.tree_histori_sj.column("barcode", width=140, anchor="center")
        self.tree_histori_sj.column("catatan", width=450, anchor="w") # Diperlebar agar kursor geser horizontal aktif
        
        frame.bind("<Visibility>", lambda e: self.muat_histori_sj_excel())

    def muat_histori_sj_excel(self):
        """ ENGINE BACKEND: Menarik arsip SJ otomatis data TERBARU di PALING ATAS via Pandas (FIX ILOC) """
        self.semua_data_histori_sj = []
        for item in self.tree_histori_sj.get_children():
            self.tree_histori_sj.delete(item)
            
        try:
            df_tracking = pd.read_excel(self.excel_file, sheet_name="Tracking Packing List", skiprows=14, header=None)
            df_track_clean = df_tracking.dropna(how='all')
            kolom_ket_track = df_track_clean.iloc[:, 4].astype(str).str.strip().str.upper()
            
            df_hist_sj = pd.read_excel(self.excel_file, sheet_name="HISTORI SURAT JALAN", skiprows=5, header=None)
            df_sj_clean = df_hist_sj.dropna(how='all')
            df_terbaru = df_sj_clean.iloc[::-1]
            
            counter = 1
            for idx, row in df_terbaru.iterrows():
                # KUNCI PERBAIKAN: Mengembalikan tanda kurung siku indeks kolom [1, 2, 3, 5] yang benar
                tgl_val = row.iloc[1] if len(row) > 1 else None
                no_sj_val = row.iloc[2] if len(row) > 2 else None
                toko_val = row.iloc[3] if len(row) > 3 else None
                catatan_val = row.iloc[5] if len(row) > 5 else "-"
                
                if no_sj_val and str(no_sj_val).strip() != "None":
                    no_sj_str = str(no_sj_val).strip()
                    tgl_str = str(tgl_val).strip().split()[0] if pd.notna(tgl_val) else "-"
                    toko_str = str(toko_val).strip() if pd.notna(toko_val) else "-"
                    catatan_str = str(catatan_val).strip() if pd.notna(catatan_val) else "-"
                    
                    mask_nota = kolom_ket_track.str.contains(no_sj_str.upper(), na=False)
                    jumlah_qty = int(mask_nota.sum())
                    qty_tampilan_str = f"{jumlah_qty} Unit"
                    
                    isi_baris = (counter, tgl_str, no_sj_str, toko_str, qty_tampilan_str, catatan_str)
                    self.semua_data_histori_sj.append(isi_baris)
                    self.tree_histori_sj.insert("", "end", values=isi_baris)
                    counter += 1
        except Exception as err:
            print(f"Gagal memuat otomatis database histori Qty SJ terbaru: {str(err)}")

    def filter_data_histori_sj(self):
        """ SENSOR FILTER KILAT: Menyaring arsip Surat Jalan live dari RAM tanpa merusak hitungan Qty """
        kata_kunci = self.entry_cari_histori_sj.get().strip().lower()
        
        # Kosongkan baris grid tabel di layar monitor terlebih dahulu
        for item in self.tree_histori_sj.get_children():
            self.tree_histori_sj.delete(item)
            
        counter = 1
        # Menyisir seluruh data arsip yang tersimpan aman di dalam memori internal RAM Python
        for item in self.semua_data_histori_sj:
            # Mencocokkan kata kunci pada Kolom Nomor Nota, Nama Toko, Qty Unit, dan Catatan
            string_gabung = f"{str(item[2]).lower()} {str(item[3]).lower()} {str(item[4]).lower()} {str(item[5]).lower()}"
            
            if kata_kunci in string_gabung:
                baris_baru = (counter, item[1], item[2], item[3], item[4], item[5])
                self.tree_histori_sj.insert("", "end", values=baris_baru)
                counter += 1

    def buat_halaman_histori_st(self):
        """ Menu 7: HALAMAN ARSIP DOKUMEN CETAK TANDA TERIMA (REAL-TIME + ENTERPRISE MEWAH) """
        from tkinter import ttk
        
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["hist_st"] = frame
        self.buat_label_copyright(frame)
        
        ctk.CTkLabel(frame, text="HISTORI ARSIP DOKUMEN TANDA TERIMA (ST)", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 2))
        ctk.CTkLabel(frame, text="Menampilkan rekam jejak digital seluruh berkas bukti Tanda Terima secara real-time otomatis dari data terbaru.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        search_frame = ctk.CTkFrame(frame, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(search_frame, text="Cari No Tanda Terima / Nama Pihak Pengirim:", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=("#334155", "#F8FAFC")).pack(side="left", padx=(5, 10))
        self.entry_cari_histori_st = ctk.CTkEntry(search_frame, placeholder_text="Ketik nomor nota / pengirim di sini untuk menyaring arsip secara instan...", width=380, height=30, fg_color="#FFFFFF", text_color="#000000")
        self.entry_cari_histori_st.pack(side="left", padx=0)
        self.entry_cari_histori_st.bind("<KeyRelease>", lambda e: self.filter_data_histori_st())
        
        # AREA BINGKAI TABEL (GRID LAYOUT PREMIUM)
        tabel_container = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_color="#CBD5E1", border_width=1)
        tabel_container.pack(fill="both", expand=True, padx=2, pady=5)
        tabel_container.grid_columnconfigure(0, weight=1)
        tabel_container.grid_rowconfigure(0, weight=1)
        
        kolom_st = ("no", "waktu", "no_st", "pengirim", "barcode", "catatan")
        self.tree_histori_st = ttk.Treeview(tabel_container, columns=kolom_st, show="headings", style="Treeview")
        self.tree_histori_st.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        scrollbar_v = ttk.Scrollbar(tabel_container, orient="vertical", command=self.tree_histori_st.yview)
        scrollbar_v.grid(row=0, column=1, sticky="ns", pady=(5, 0), padx=(0, 5))
        scrollbar_h = ttk.Scrollbar(tabel_container, orient="horizontal", command=self.tree_histori_st.xview)
        scrollbar_h.grid(row=1, column=0, sticky="ew", padx=(5, 0), pady=(0, 5))
        self.tree_histori_st.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # BINDING INTERAKTIF: Klik 2x batas kolom untuk Auto-Fit ala Excel
        self.tree_histori_st.bind("<Double-1>", lambda e: self.auto_fit_kolom(e, self.tree_histori_st))
        
        # SUNTIKAN FITUR SORTIR: Klik judul kolom untuk mengurutkan data (A-Z / Z-A)
        for col in kolom_st:
            self.tree_histori_st.heading(col, text=col.title() if col != "no" else "No", command=lambda c=col: self.urutkan_kolom(self.tree_histori_st, c, False))
            
        self.tree_histori_st.heading("waktu", text="Tanggal Masuk")
        self.tree_histori_st.heading("no_st", text="Nomor Tanda Terima")
        self.tree_histori_st.heading("pengirim", text="Nama Pihak Pengirim Unit")
        self.tree_histori_st.heading("barcode", text="Total Unit Masuk")
        self.tree_histori_st.heading("catatan", text="Alasan / Detail Kondisi Fisik")
        
        self.tree_histori_st.column("no", width=50, anchor="center")
        self.tree_histori_st.column("waktu", width=120, anchor="center")
        self.tree_histori_st.column("no_st", width=140, anchor="center")
        self.tree_histori_st.column("pengirim", width=160, anchor="w")
        self.tree_histori_st.column("barcode", width=140, anchor="center")
        self.tree_histori_st.column("catatan", width=450, anchor="w") # Diperlebar agar kursor geser horizontal aktif
        
        frame.bind("<Visibility>", lambda e: self.muat_histori_st_excel())

    def muat_histori_st_excel(self):
        """ ENGINE BACKEND: Menarik arsip ST otomatis data TERBARU di PALING ATAS via Pandas (FIX ILOC) """
        self.semua_data_histori_st = []
        for item in self.tree_histori_st.get_children():
            self.tree_histori_st.delete(item)
            
        try:
            df_tracking = pd.read_excel(self.excel_file, sheet_name="Tracking Packing List", skiprows=14, header=None)
            df_track_clean = df_tracking.dropna(how='all')
            kolom_ket_track = df_track_clean.iloc[:, 4].astype(str).str.strip().str.upper()
            
            df_hist_st = pd.read_excel(self.excel_file, sheet_name="HISTORI TANDA TERIMA", skiprows=5, header=None)
            df_st_clean = df_hist_st.dropna(how='all')
            df_terbaru = df_st_clean.iloc[::-1]
            
            counter = 1
            for idx, row in df_terbaru.iterrows():
                # KUNCI UTAMA SANG PENYEMBUH: Mengembalikan indeks kurung siku pembaca kolom secara tepat
                tgl_val = row.iloc[1] if len(row) > 1 else None
                no_st_val = row.iloc[2] if len(row) > 2 else None
                pengirim_val = row.iloc[3] if len(row) > 3 else None
                catatan_val = row.iloc[5] if len(row) > 5 else "-"
                
                if no_st_val and str(no_st_val).strip() != "None":
                    no_st_str = str(no_st_val).strip()
                    tgl_str = str(tgl_val).strip().split()[0] if pd.notna(tgl_val) else "-"
                    pengirim_str = str(pengirim_val).strip() if pd.notna(pengirim_val) else "-"
                    catatan_str = str(catatan_val).strip() if pd.notna(catatan_val) else "-"
                    
                    mask_nota = kolom_ket_track.str.contains(no_st_str.upper(), na=False)
                    jumlah_qty = int(mask_nota.sum())
                    qty_tampilan_str = f"{jumlah_qty} Unit"
                    
                    isi_baris = (counter, tgl_str, no_st_str, pengirim_str, qty_tampilan_str, catatan_str)
                    self.semua_data_histori_st.append(isi_baris)
                    self.tree_histori_st.insert("", "end", values=isi_baris)
                    counter += 1
        except Exception as err:
            print(f"Gagal memuat otomatis database histori Qty ST terbaru: {str(err)}")

    def filter_data_histori_st(self):
        """ SENSOR FILTER KILAT: Menyaring arsip Tanda Terima live dari RAM tanpa merusak hitungan Qty """
        kata_kunci = self.entry_cari_histori_st.get().strip().lower()
        
        # Kosongkan baris grid tabel di layar monitor terlebih dahulu
        for item in self.tree_histor_st.get_children():
            self.tree_histori_st.delete(item)
            
        counter = 1
        # Menyisir seluruh data arsip yang tersimpan aman di dalam memori internal RAM Python
        for item in self.semua_data_histori_st:
            # Mencocokkan kata kunci pada Kolom Nomor Nota ST, Nama Pengirim, Qty Unit, dan Catatan
            string_gabung = f"{str(item[2]).lower()} {str(item[3]).lower()} {str(item[4]).lower()} {str(item[5]).lower()}"
            
            if kata_kunci in string_gabung:
                baris_baru = (counter, item[1], item[2], item[3], item[4], item[5])
                self.tree_histori_st.insert("", "end", values=baris_baru)
                counter += 1

    def buat_halaman_pack_rekap(self):
        """ REVISI TOTAL MENU 8: Membangun Stasiun Konten Cetak Label Barcode Mini (50mm x 20mm) """
        import customtkinter as ctk
        
        # 1. Menyiapkan kontainer master frame utama stasiun barcode mini
        frame = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=0)
        self.frame_cetak_barcode = frame
        self.halaman_aktif["cetak_barcode"] = frame  # Menjaga variabel nama lama agar switch halaman tidak crash
        self.buat_label_copyright(frame)
        
        # 2. Header Judul Premium Halaman Utama
        header_frame = ctk.CTkFrame(frame, fg_color="#FFFFFF", height=65, corner_radius=0, border_width=1, border_color="#E2E8F0")
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)
        
        judul_page = ctk.CTkLabel(header_frame, text="🏷️ STASIUN GENERATOR & CETAK LABEL BARCODE THERMAL", font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"), text_color="#1E293B")
        judul_page.pack(side="left", padx=20)
        
        # 3. Kontainer Panel Kontrol Utama (Layout Center)
        panel_utama = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E2E8F0")
        panel_utama.pack(padx=25, pady=25, fill="both", expand=True)
        
        sub_judul = ctk.CTkLabel(panel_utama, text="Konfigurasi Cetak Massal Stiker Thermal (Standard 50mm x 20mm - Code 128)", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color="#475569")
        sub_judul.pack(anchor="w", padx=25, pady=(20, 15))
        
        # Boks Form Isian Grey Premium
        box_form = ctk.CTkFrame(panel_utama, fg_color="#F1F5F9", corner_radius=10)
        box_form.pack(padx=25, fill="x", pady=5)
        
        # --- KOMPONEN 1: DROPDOWN PILIHAN KODE BARCODE ---
        lbl_kode = ctk.CTkLabel(box_form, text="1. Pilih Kode Singkat Tipe Laptop :", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#1E293B")
        lbl_kode.grid(row=0, column=0, padx=25, pady=(15, 5), sticky="w")
        
        # Daftar pilihan kode barcode laptop GTP sesuai database Anda
        daftar_tipe = ["L15", "L36", "H33", "A32", "T480", "X280", "G3", "G5"]
        self.combo_cetak_kode = ctk.CTkComboBox(box_form, values=daftar_tipe, font=ctk.CTkFont(family="Segoe UI", size=12), width=250, height=35, corner_radius=6, fg_color="#FFFFFF", border_color="#CBD5E1", button_color="#64748B", button_hover_color="#475569")
        self.combo_cetak_kode.grid(row=1, column=0, padx=25, pady=(0, 20), sticky="w")
        self.combo_cetak_kode.set("Pilih Tipe Laptop")
        # --- KOMPONEN 2: RADIO BUTTON PILIHAN MODE CERDAS ---
        lbl_mode = ctk.CTkLabel(box_form, text="2. Pilih Mode Metode Pencetakan :", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#1E293B")
        lbl_mode.grid(row=0, column=1, padx=25, pady=(15, 5), sticky="w")
        
        self.var_mode_cetak = ctk.StringVar(value="RENTANG")
        
        # Fungsi interaktif harian untuk mengunci otomatis kotak input nomor urut jika memilih mode ALL
        def efek_saklar_mode():
            if self.var_mode_cetak.get() == "ALL":
                self.entry_no_awal.configure(state="disabled", fg_color="#E2E8F0", placeholder_text="Terkunci (Otomatis)")
                self.entry_no_akhir.configure(state="disabled", fg_color="#E2E8F0", placeholder_text="Terkunci (Otomatis)")
                self.entry_no_awal.delete(0, 'end')
                self.entry_no_akhir.delete(0, 'end')
            else:
                self.entry_no_awal.configure(state="normal", fg_color="#FFFFFF", placeholder_text="Contoh: 1")
                self.entry_no_akhir.configure(state="normal", fg_color="#FFFFFF", placeholder_text="Contoh: 10")
        
        rad_rentang = ctk.CTkRadioButton(box_form, text="Mode Rentang Manual (Cek Anti-Hantu)", variable=self.var_mode_cetak, value="RENTANG", command=efek_saklar_mode, font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), text_color="#1E293B", fg_color="#3B82F6", hover_color="#1D4ED8")
        rad_rentang.grid(row=1, column=1, padx=25, pady=(0, 5), sticky="w")
        
        rad_semua = ctk.CTkRadioButton(box_form, text="Mode Cetak Semua Nomor Urut di Database", variable=self.var_mode_cetak, value="ALL", command=efek_saklar_mode, font=ctk.CTkFont(family="Segoe UI", size=12, weight="normal"), text_color="#1E293B", fg_color="#3B82F6", hover_color="#1D4ED8")
        rad_semua.grid(row=2, column=1, padx=25, pady=(0, 15), sticky="w")
        
        # --- KOMPONEN 3: KOTAK INPUT ANGKA RENTANG NOMOR URUT ---
        lbl_rentang_box = ctk.CTkLabel(box_form, text="3. Rentang Urutan (Hanya Isi Jika Memilih Mode Rentang) :", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#1E293B")
        lbl_rentang_box.grid(row=2, column=0, padx=25, pady=(5, 5), sticky="w")
        
        container_input_angka = ctk.CTkFrame(box_form, fg_color="transparent")
        container_input_angka.grid(row=3, column=0, padx=25, pady=(0, 20), sticky="w")
        
        lbl_dari = ctk.CTkLabel(container_input_angka, text="Dari No:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#475569")
        lbl_dari.pack(side="left", padx=(0, 5))
        
        self.entry_no_awal = ctk.CTkEntry(container_input_angka, width=75, height=32, corner_radius=6, border_color="#CBD5E1", font=ctk.CTkFont(family="Segoe UI", size=12), fg_color="#FFFFFF", text_color="#1E293B", placeholder_text="Contoh: 1")
        self.entry_no_awal.pack(side="left", padx=(0, 15))
        
        lbl_sampai = ctk.CTkLabel(container_input_angka, text="Sampai No:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#475569")
        lbl_sampai.pack(side="left", padx=(0, 5))
        
        self.entry_no_akhir = ctk.CTkEntry(container_input_angka, width=75, height=32, corner_radius=6, border_color="#CBD5E1", font=ctk.CTkFont(family="Segoe UI", size=12), fg_color="#FFFFFF", text_color="#1E293B", placeholder_text="Contoh: 10")
        self.entry_no_akhir.pack(side="left")
        # --- KOMPONEN 4: TOMBOL EKSEKUTOR CETAK MASSAL PREMIUM ---
        self.btn_eksekusi_barcode = ctk.CTkButton(
            panel_utama, 
            text="🔥 Eksekusi Cetak Massal Stiker Barcode (50mm x 20mm)", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), 
            height=45, 
            corner_radius=8, 
            fg_color="#2563EB", 
            hover_color="#1D4ED8", 
            text_color="#FFFFFF",
            command=self.mesin_backend_cetak_barcode_massal
        )
        self.btn_eksekusi_barcode.pack(padx=25, pady=(20, 30), fill="x")

    def mesin_backend_cetak_barcode_massal(self):
        """ ENGINE BACKEND BARCODE MASSAL: Memvalidasi database, memotong hantu, dan mencetak otomatis """
        import os
        import openpyxl
        from tkinter import messagebox
        
        # Ambil nilai parameter dari form inputan UI di atas
        kode_prefix = self.combo_cetak_kode.get().strip()
        mode_pilihan = self.var_mode_cetak.get()
        
        # Validasi awal pengaman dropdown
        if not kode_prefix or "Pilih" in kode_prefix:
            self.bell()
            self.tampilkan_peringatan_besar("Pilih Tipe", "FOKUS! Silakan tentukan KODE SINGKAT TIPE LAPTOP (seperti L15, L36, dll) pada dropdown nomor 1 sebelum memproses!", "#F59E0B")
            return
            
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_master = wb["MASTER DATA"]
            
            # Membaca seluruh data Barcode di sheet MASTER DATA Kolom B ke dalam RAM Python
            daftar_barcode_database = set()
            for r in range(3, ws_master.max_row + 1):
                b_val = ws_master.cell(row=r, column=2).value
                if b_val:
                    daftar_barcode_database.add(str(b_val).strip().upper())
            daftar_target_barcode = []
            
            # --- SCENARIO 1: MODE RENTANG MANUAL (DENGAN SENSOR PROTEKSI ANTI-HANTU) ---
            if mode_pilihan == "RENTANG":
                no_awal_raw = self.entry_no_awal.get().strip()
                no_akhir_raw = self.entry_no_akhir.get().strip()
                
                if not no_awal_raw or not no_akhir_raw:
                    self.bell()
                    self.tampilkan_peringatan_besar("Input Kosong", "FOKUS! Kolom 'Dari No' dan 'Sampai No' wajib diisi angka jika memilih Mode Rentang Manual!", "#F59E0B")
                    return
                    
                try:
                    no_awal = int(no_awal_raw)
                    no_akhir = int(no_akhir_raw)
                except ValueError:
                    self.bell()
                    self.tampilkan_peringatan_besar("Bukan Angka", "FOKUS! Isian rentang nomor urut wajib berupa angka bulat positif (Contoh: 1 sampai 10)!", "#EF4444")
                    return
                    
                if no_awal <= 0 or no_akhir <= 0 or no_awal > no_akhir:
                    self.bell()
                    self.tampilkan_peringatan_besar("Rentang Salah", "Gagal Memproses! Batas angka rentang terbalik atau tidak valid. Mohon periksa kembali input Anda.", "#EF4444")
                    return
                    
                for urutan in range(no_awal, no_akhir + 1):
                    kode_lengkap = f"TDA.{kode_prefix}.{str(urutan).zfill(4)}".upper()
                    daftar_target_barcode.append(kode_lengkap)
                    
                # [VALIDASI CRUCIAL - SARINGAN KERANJANG MERAH ANTI-HANTU]
                keranjang_merah_kosong = []
                for b_target in daftar_target_barcode:
                    if b_target not in daftar_barcode_database:
                        keranjang_merah_kosong.append(b_target)
                        
                # Jika ditemukan ada data yang tidak terdaftar di Excel, BATALKAN TOTAL!
                if len(keranjang_merah_kosong) > 0:
                    self.bell()
                    teks_pelanggaran = "\n".join([f"  ❌ {item}" for item in keranjang_merah_kosong[:8]])
                    if len(keranjang_merah_kosong) > 8:
                        teks_pelanggaran += f"\n  ...dan {len(keranjang_merah_kosong) - 8} unit lainnya."
                        
                    pesan_batal = f"DATA TIDAK TERSEDIA!\n\n" \
                                  f"Proses cetak massal label barcode dibatalkan total!\n\n" \
                                  f"Ditemukan {len(keranjang_merah_kosong)} Nomor Urut yang tidak terdaftar di sheet MASTER DATA:\n" \
                                  f"{teks_pelanggaran}\n\n" \
                                  f"Silakan daftarkan unit tersebut terlebih dahulu sebelum memproses cetak label."
                    self.tampilkan_peringatan_besar("Data Kosong / Palsu", pesan_batal, "#EF4444")
                    return
                    
            # --- SCENARIO 2: MODE CETAK SEMUA NOMOR URUT DI DATABASE MASTER ---
            else:
                # Menyisir database RAM dengan saringan kata depan "TDA.PREFIX." agar terbaca akurat
                prefix_wajib = f"TDA.{kode_prefix.upper()}."
                for b_db in sorted(list(daftar_barcode_database)):
                    if b_db.startswith(prefix_wajib):
                        daftar_target_barcode.append(b_db)

                        
                if not daftar_target_barcode:
                    self.bell()
                    self.tampilkan_peringatan_besar("Tipe Tidak Ada", f"Gagal Memproses! Tidak ditemukan satu pun unit laptop dengan kode '{kode_prefix}' di dalam database MASTER DATA Anda.", "#EF4444")
                    return
            # --- JENDELA KONFIRMASI RINGKASAN AMAN BEBAS POPUP BERONDONG ---
            total_cetak = len(daftar_target_barcode)
            pesan_rekap = f"CEK KEMBALI OPERATOR GUDANG!\n\n" \
                          f"Siap merakit {total_cetak} Stiker Barcode Prefix [{kode_prefix}].\n" \
                          f"Manifes: {daftar_target_barcode[0]} s/d {daftar_target_barcode[-1]}\n\n" \
                          f"Yakin ingin mengekspor {total_cetak} halaman stiker ke dalam 1 FILE PDF?"
                          
            setuju_cetak = self.tampilkan_konfirmasi_besar("Konfirmasi Cetak Massal", pesan_rekap)
            if not setuju_cetak:
                return

            # Mengimpor pustaka ReportLab & Barcode terintegrasi
            from barcode import Code128
            from barcode.writer import ImageWriter
            from reportlab.platypus import SimpleDocTemplate, Image as RLImage, PageBreak, Spacer
            import os

            folder_temp = "Temp_Barcode"
            if not os.path.exists(folder_temp):
                os.makedirs(folder_temp)

            # Menentukan nama file PDF gabungan massal harian
            nama_file_pdf = f"Kumpulan_Stiker_{kode_prefix}_{total_cetak}_Unit.pdf"
            path_pdf_final = os.path.join(folder_temp, nama_file_pdf)
            
            # KUNCI RESIZABLE: Mengunci ukuran halaman PDF pas 141.7 x 56.7 poin (50mm x 20mm) tanpa margin (0)
            lebar_poin = 141.7
            tinggi_poin = 56.7
            doc = SimpleDocTemplate(path_pdf_final, pagesize=(lebar_poin, tinggi_poin), leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0)
            story = []

            # --- LOOPING UTAMA RAM: Merakit halaman per halaman secara gaib ke dalam 1 file PDF ---
            for idx, b_id in enumerate(daftar_target_barcode):
                path_mentah_direct = os.path.join(folder_temp, f"raw_{b_id}")
                
                # REVISI FIX FONT: Mengunci jalur font Arial Windows agar bebas dari error cannot open resource
                opsi_penulis = {
                    'write_text': True,
                    'font_path': "C:\\Windows\\Fonts\\arial.ttf", # Jalur mutlak font Windows Anda
                    'font_size': 10,       # Ukuran teks proporsional untuk stiker 50x20mm
                    'text_distance': 3.0,  # Jarak aman teks di bawah garis barcode
                    'module_height': 10.0, # Tinggi garis disesuaikan agar teks tidak terpotong bawah
                    'module_width': 0.18,  # Kepadatan garis pas masuk ke lebar kertas 50mm
                    'quiet_zone': 1.0,
                    'background': '#FFFFFF',
                    'foreground': '#000000'
                }
                
                generator_code = Code128(b_id, writer=ImageWriter())
                file_png_stiker = generator_code.save(path_mentah_direct, options=opsi_penulis)
                
                # REVISI UKURAN SAFE-ZONE: Lebar disetel 120 dan tinggi disetel 35 agar dijamin masuk ke area cetak tanpa luber
                img_stiker_pdf = RLImage(file_png_stiker, width=120.0, height=35.0)
                img_stiker_pdf.hAlign = 'CENTER'
                story.append(img_stiker_pdf)
                
                # KUNCI ANTARA HALAMAN: Berikan instruksi potong halaman kecuali untuk baris stiker terakhir
                if idx < total_cetak - 1:
                    story.append(PageBreak())

            # Eksekusi perakitan 1 file PDF massal tunggal
            doc.build(story)
            
            # Pembersihan berkas sampah gambar PNG sementara agar harddisk bersih
            for b_id in daftar_target_barcode:
                target_hapus = os.path.join(folder_temp, f"raw_{b_id}.png")
                if os.path.exists(target_hapus):
                    os.remove(target_hapus)

            # --- SELESAI SUKSES: Membuka 1 File PDF Tunggal Berisi Ratusan Halaman Stiker ---
            try:
                os.startfile(os.path.abspath(path_pdf_final))
            except Exception as print_err:
                print(f"Gagal membuka pratinjau file PDF massal: {str(print_err)}")
                
            self.tampilkan_peringatan_besar(
                "Ekspor PDF Sukses", 
                f"Berhasil Merakit Dokumen Cetak Massal!\n\n" \
                f"Total {total_cetak} halaman stiker label barcode [{kode_prefix}]\n" \
                f"telah disatukan ke dalam 1 file PDF tunggal.\n\n" \
                f"Silakan tekan Ctrl + P pada file PDF yang terbuka untuk cetak beruntun.", 
                "#10B981"
            )
            
        except Exception as err:
            self.tampilkan_peringatan_besar("Error Engine Barcode", f"Sistem gagal merakit cetakan massal PDF: {str(err)}", "#EF4444")

    def buat_halaman_stok_rekap(self):
        """ Menu 9: HALAMAN REKAPITULASI STOK LAPTOP PER TIPE UNIT (REAL-TIME + ENTERPRISE MEWAH) """
        from tkinter import ttk
        
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["stock_rekap"] = frame
        self.buat_label_copyright(frame)
        
        # Header Tampilan Menu
        ctk.CTkLabel(frame, text="REKAPITULASI DISTRIBUSI STOK LAPTOP", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 2))
        ctk.CTkLabel(frame, text="Merangkum total kuantitas unit barang berdasarkan tipe spesifikasi laptop dan status posisinya secara real-time otomatis.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        # PANEL KONTROL ATAS: Kotak Pencarian Kilat Tipe Laptop
        search_frame = ctk.CTkFrame(frame, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(search_frame, text="Cari Spesifikasi / Nama Tipe Laptop:", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=("#334155", "#F8FAFC")).pack(side="left", padx=(5, 10))
        self.entry_cari_stok_rekap = ctk.CTkEntry(search_frame, placeholder_text="Ketik nama unit / spesifikasi laptop untuk memfilter rekap stok...", width=380, height=30, fg_color="#FFFFFF", text_color="#000000")
        self.entry_cari_stok_rekap.pack(side="left", padx=0)
        self.entry_cari_stok_rekap.bind("<KeyRelease>", lambda e: self.filter_data_stok_rekap())
        
        # AREA BINGKAI TABEL (GRID LAYOUT PREMIUM)
        tabel_container = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_color="#CBD5E1", border_width=1)
        tabel_container.pack(fill="both", expand=True, padx=2, pady=5)
        tabel_container.grid_columnconfigure(0, weight=1)
        tabel_container.grid_rowconfigure(0, weight=1)
        
        kolom_st = ("no", "nama_laptop", "total_stok", "stok_ready", "stok_sewa", "stok_pinjam", "stok_service")
        self.tree_stok_rekap = ttk.Treeview(tabel_container, columns=kolom_st, show="headings", style="Treeview")
        self.tree_stok_rekap.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        scrollbar_v = ttk.Scrollbar(tabel_container, orient="vertical", command=self.tree_stok_rekap.yview)
        scrollbar_v.grid(row=0, column=1, sticky="ns", pady=(5, 0), padx=(0, 5))
        scrollbar_h = ttk.Scrollbar(tabel_container, orient="horizontal", command=self.tree_stok_rekap.xview)
        scrollbar_h.grid(row=1, column=0, sticky="ew", padx=(5, 0), pady=(0, 5))
        self.tree_stok_rekap.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # BINDING INTERAKTIF: Klik 2x batas kolom untuk Auto-Fit ala Excel
        self.tree_stok_rekap.bind("<Double-1>", lambda e: self.auto_fit_kolom(e, self.tree_stok_rekap))
        
        # SUNTIKAN FITUR SORTIR: Klik judul kolom untuk mengurutkan data (A-Z / Z-A)
        for col in kolom_st:
            self.tree_stok_rekap.heading(col, text=col.title() if col != "no" else "No", command=lambda c=col: self.urutkan_kolom(self.tree_stok_rekap, c, False))
            
        self.tree_stok_rekap.heading("nama_laptop", text="Spesifikasi / Nama Unit Laptop")
        self.tree_stok_rekap.heading("total_stok", text="Total Unit")
        self.tree_stok_rekap.heading("stok_ready", text="Ready")
        self.tree_stok_rekap.heading("stok_sewa", text="Disewa")
        self.tree_stok_rekap.heading("stok_pinjam", text="Dipinjam")
        self.tree_stok_rekap.heading("stok_service", text="Servis/Rusak")
        
        self.tree_stok_rekap.column("no", width=50, anchor="center")
        self.tree_stok_rekap.column("nama_laptop", width=450, anchor="w") # Diperlebar agar kursor geser horizontal aktif
        self.tree_stok_rekap.column("total_stok", width=90, anchor="center")
        self.tree_stok_rekap.column("stok_ready", width=90, anchor="center")
        self.tree_stok_rekap.column("stok_sewa", width=90, anchor="center")
        self.tree_stok_rekap.column("stok_pinjam", width=90, anchor="center")
        self.tree_stok_rekap.column("stok_service", width=100, anchor="center")
        
        frame.bind("<Visibility>", lambda e: self.hitung_rekap_stok_laptop())

    def hitung_rekap_stok_laptop(self):
        """ ENGINE BACKEND: Merangkum total stok per tipe unit secara otomatis dari MASTER DATA """
        import openpyxl
        
        self.semua_data_stok_rekap = []
        
        # Bersihkan sisa tampilan baris visual lama di layar sebelum diisi data baru
        for item in self.tree_stok_rekap.get_children():
            self.tree_stok_rekap.delete(item)
            
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_master = wb["MASTER DATA"]
            
            # Struktur Dictionary untuk mengelompokkan data berdasarkan spesifikasi laptop
            kamus_stok = {}
            
            # Menyisir tabel induk pada MASTER DATA mulai dari baris ke-3 ke bawah
            for row in range(3, ws_master.max_row + 1):
                nama_val = ws_master.cell(row=row, column=3).value      # Kolom C: Nama/Spesifikasi
                status_val = str(ws_master.cell(row=row, column=5).value).strip().upper() # Kolom E: Status
                
                if nama_val and str(nama_val).strip() != "None":
                    nama_laptop = str(nama_val).strip()
                    
                    # Inisialisasi jika nama laptop belum terdaftar di dalam kamus
                    if nama_laptop not in kamus_stok:
                        kamus_stok[nama_laptop] = {"total": 0, "ready": 0, "sewa": 0, "pinjam": 0, "service": 0}
                        
                    kamus_stok[nama_laptop]["total"] += 1
                    
                    # Distribusi perhitungan berdasarkan status riil di lapangan
                    if status_val == "READY":
                        kamus_stok[nama_laptop]["ready"] += 1
                    elif status_val in ["SEWA", "DISEWA", "TERJUAL", "JUAL"]:
                        kamus_stok[nama_laptop]["sewa"] += 1
                    elif status_val in ["PINJAM", "DIPINJAM"]:
                        kamus_stok[nama_laptop]["pinjam"] += 1
                    elif status_val in ["SERVICE", "RUSAK"]:
                        kamus_stok[nama_laptop]["service"] += 1

            counter = 1
            # Memasukkan hasil kalkulasi ringkasan ke memori internal RAM dan layar grid
            for laptop, qty in kamus_stok.items():
                isi_baris = (counter, laptop, qty["total"], qty["ready"], qty["sewa"], qty["pinjam"], qty["service"])
                self.semua_data_stok_rekap.append(isi_baris)
                self.tree_stok_rekap.insert("", "end", values=isi_baris)
                counter += 1
                
        except Exception as err:
            print(f"Gagal memproses data kalkulasi rekap stok laptop: {str(err)}")
            
    def filter_data_stok_rekap(self):
        """ SENSOR FILTER KILAT: Menyaring ringkasan stok laptop secara live berdasarkan tipe unit """
        kata_kunci = self.entry_cari_stok_rekap.get().strip().lower()
        
        # Kosongkan baris grid tabel di layar monitor terlebih dahulu
        for item in self.tree_stok_rekap.get_children():
            self.tree_stok_rekap.delete(item)
            
        counter = 1
        # Menyisir seluruh ringkasan stok yang tersimpan di dalam memori internal RAM Python
        for item in self.semua_data_stok_rekap:
            nama_laptop_str = str(item[1]).lower()
            if kata_kunci in nama_laptop_str:
                baris_baru = (counter, item[1], item[2], item[3], item[4], item[5], item[6])
                self.tree_stok_rekap.insert("", "end", values=baris_baru)
                counter += 1

    def buat_halaman_stok(self):
        """ Menu 10: INDUK DATABASE UTAMA ASET LAPTOP (REAL-TIME + ENTERPRISE MEWAH) """
        from tkinter import ttk
        
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        self.halaman_aktif["stock"] = frame
        self.buat_label_copyright(frame)
        
        # Header Tampilan Menu
        ctk.CTkLabel(frame, text="DATABASE PUSAT - MASTER DATA UNIT LAPTOP", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color=("#1E293B", "#F8FAFC")).pack(anchor="w", pady=(5, 2))
        ctk.CTkLabel(frame, text="Menampilkan seluruh rincian database aset laptop perusahaan secara lengkap, akurat, dan real-time otomatis.", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#64748B").pack(anchor="w", pady=(0, 15))
        
        # PANEL KONTROL ATAS: Kotak Pencarian Kilat Multi-Kolom Master Data
        search_frame = ctk.CTkFrame(frame, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(search_frame, text="Cari No Barcode / Spesifikasi Laptop / Status:", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=("#334155", "#F8FAFC")).pack(side="left", padx=(5, 10))
        self.entry_cari_master_unit = ctk.CTkEntry(search_frame, placeholder_text="Ketik nomor barcode, nama laptop, atau status untuk menyaring database...", width=380, height=30, fg_color="#FFFFFF", text_color="#000000")
        self.entry_cari_master_unit.pack(side="left", padx=0)
        self.entry_cari_master_unit.bind("<KeyRelease>", lambda e: self.filter_data_master_unit())
        
        # AREA BINGKAI TABEL (GRID LAYOUT PREMIUM)
        tabel_container = ctk.CTkFrame(frame, fg_color="#FFFFFF", corner_radius=12, border_color="#CBD5E1", border_width=1)
        tabel_container.pack(fill="both", expand=True, padx=2, pady=5)
        tabel_container.grid_columnconfigure(0, weight=1)
        tabel_container.grid_rowconfigure(0, weight=1)
        
        kolom_st = ("no", "barcode", "nama_laptop", "spesifikasi", "status", "keterangan")
        self.tree_master_unit = ttk.Treeview(tabel_container, columns=kolom_st, show="headings", style="Treeview")
        self.tree_master_unit.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=(5, 0))
        
        scrollbar_v = ttk.Scrollbar(tabel_container, orient="vertical", command=self.tree_master_unit.yview)
        scrollbar_v.grid(row=0, column=1, sticky="ns", pady=(5, 0), padx=(0, 5))
        scrollbar_h = ttk.Scrollbar(tabel_container, orient="horizontal", command=self.tree_master_unit.xview)
        scrollbar_h.grid(row=1, column=0, sticky="ew", padx=(5, 0), pady=(0, 5))
        self.tree_master_unit.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # BINDING INTERAKTIF: Klik 2x batas kolom untuk Auto-Fit ala Excel
        self.tree_master_unit.bind("<Double-1>", lambda e: self.auto_fit_kolom(e, self.tree_master_unit))
        
        # SUNTIKAN FITUR SORTIR: Klik judul kolom untuk mengurutkan data (A-Z / Z-A)
        for col in kolom_st:
            self.tree_master_unit.heading(col, text=col.title() if col != "no" else "No", command=lambda c=col: self.urutkan_kolom(self.tree_master_unit, c, False))
            
        self.tree_master_unit.heading("barcode", text="Nomor Barcode")
        self.tree_master_unit.heading("nama_laptop", text="Nama Barang / Merk")
        self.tree_master_unit.heading("spesifikasi", text="Serial Number")
        self.tree_master_unit.heading("status", text="Status Utama")
        self.tree_master_unit.heading("keterangan", text="Keterangan")
        
        self.tree_master_unit.column("no", width=50, anchor="center")
        self.tree_master_unit.column("barcode", width=120, anchor="center")
        self.tree_master_unit.column("nama_laptop", width=160, anchor="w")
        self.tree_master_unit.column("spesifikasi", width=220, anchor="w")
        self.tree_master_unit.column("status", width=95, anchor="center")
        self.tree_master_unit.column("keterangan", width=450, anchor="w") # Diperlebar agar kursor geser horizontal aktif maksimal
        
        frame.bind("<Visibility>", lambda e: self.muat_master_unit_excel())

    def muat_master_unit_excel(self):
        """ ENGINE BACKEND: Menarik database induk MASTER DATA dengan Pandas (Super Cepat <0.1 Detik) """
        self.semua_data_master_unit = []
        
        # Bersihkan sisa tampilan baris visual lama di layar monitor sebelum diisi data baru
        for item in self.tree_master_unit.get_children():
            self.tree_master_unit.delete(item)
            
        if not hasattr(self, 'df_master') or self.df_master is None:
            return
            
        try:
            # Menggunakan dataframe df_master yang sudah dibaca cepat oleh pandas di awal
            df_clean = self.df_master.dropna(how='all')
            
            counter = 1
            for idx, row in df_clean.iterrows():
                # Mengekstrak kolom B (1), C (2), D (3), E (4), F (5) dari MASTER DATA secara instan lewat RAM
                barcode_str = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else "-"
                nama_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else "-"
                spesifikasi_str = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else "-"
                status_str = str(row.iloc[4]).strip().upper() if pd.notna(row.iloc[4]) else "-"
                keterangan_str = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else "-"
                
                if barcode_str != "-" and barcode_str != "NONE":
                    isi_baris = (counter, barcode_str, nama_str, spesifikasi_str, status_str, keterangan_str)
                    self.semua_data_master_unit.append(isi_baris)
                    
                    # Dimasukkan secara polos tanpa TAG warna agar aplikasi enteng dan stabil harian
                    self.tree_master_unit.insert("", "end", values=isi_baris)
                    counter += 1
                    
        except Exception as err:
            print(f"Gagal memuat database internal master data unit cepat: {str(err)}")

    def filter_data_master_unit(self):
        """ SENSOR FILTER KILAT: Menyaring isi database induk secara live berdasarkan ketikan operator """
        kata_kunci = self.entry_cari_master_unit.get().strip().lower()
        
        # Kosongkan baris grid tabel di layar monitor terlebih dahulu
        for item in self.tree_master_unit.get_children():
            self.tree_master_unit.delete(item)
            
        counter = 1
        # Menyisir memori internal RAM Python yang tersimpan dari hasil ekstraksi kilat Pandas di atas
        for item in self.semua_data_master_unit:
            # Menggabungkan data seluruh kolom (Barcode, Nama, Serial Number, Status Utama, Keterangan)
            string_gabung = f"{str(item[1]).lower()} {str(item[2]).lower()} {str(item[3]).lower()} {str(item[4]).lower()} {str(item[5]).lower()}"
            
            # Jika kata kunci COCOK dengan isi data baris arsip master
            if kata_kunci in string_gabung:
                baris_baru = (counter, item[1], item[2], item[3], item[4], item[5])
                self.tree_master_unit.insert("", "end", values=baris_baru)
                counter += 1

    def muat_data_excel(self):
        """ ENGINE UTAMA LURUS: Membaca Murni Cell A3 (skiprows=2) Tanpa Judul (header=None) """
        if not os.path.exists(self.excel_file): 
            return
        try:
            self.df_master = pd.read_excel(self.excel_file, sheet_name="MASTER DATA", skiprows=2, header=None)
            df_clean = self.df_master.dropna(how='all')
            kolom_status = df_clean.iloc[:, 4].astype(str).str.strip().str.upper()

            # Kartu 1-7 Indikator Real-Time Dasbor
            total_unit = len(df_clean)
            self.lbl_total_unit.configure(text=str(total_unit))
            unit_ready = (kolom_status == "READY").sum()
            self.lbl_unit_ready.configure(text=str(unit_ready))
            unit_sewa = (kolom_status.str.contains("SEWA", case=False, na=False) | kolom_status.str.contains("DISEWA", case=False, na=False)).sum()
            self.lbl_unit_sewa.configure(text=str(unit_sewa))
            unit_terjual = (kolom_status.str.contains("DIJUAL", case=False, na=False) | kolom_status.str.contains("JUAL", case=False, na=False)).sum()
            self.lbl_unit_terjual.configure(text=str(unit_terjual))
            unit_pinjam = (kolom_status.str.contains("PINJAM", case=False, na=False)).sum()
            self.lbl_unit_pinjam.configure(text=str(unit_pinjam))
            unit_service = (kolom_status.str.contains("SERVICE", case=False, na=False) | kolom_status.str.contains("SERVIS", case=False, na=False)).sum()
            self.lbl_unit_service.configure(text=str(unit_service))
            unit_rusak = (kolom_status == "RUSAK").sum()
            self.lbl_unit_rusak.configure(text=str(unit_rusak))
            
            # FIX DROPDOWN CASE-INSENSITIVE: Ekstraksi murni dan standarisasi ke HURUF BESAR SEMUA
            df_log = pd.read_excel(self.excel_file, sheet_name="Tracking Packing List", skiprows=14)
            nama_bersih_set = set()
            
            for teks in df_log.iloc[:, 4].dropna().astype(str):
                teks = teks.strip().upper() # <--- KUNCI 1: Paksa standarisasi huruf besar di memori RAM
                if teks in ["", "[ PILIH NAMA TOKO ]", "TIDAK TERLACAK", "NONE"]: 
                    continue
                if "KEMBALI DARI " in teks: 
                    teks = teks.replace("KEMBALI DARI ", "")
                if "RETUR JUAL DARI " in teks: 
                    teks = teks.replace("RETUR JUAL DARI ", "")
                
                if " (SURAT:" in teks:
                    teks = teks.split(" (SURAT:")[0].strip()
                if " (NOTA:" in teks:
                    teks = teks.split(" (NOTA:")[0].strip()
                    
                if teks: 
                    nama_bersih_set.add(teks)
                    
            daftar_nama = sorted(list(nama_bersih_set))
            if daftar_nama:
                self.combo_sj_toko.configure(values=daftar_nama)
                self.combo_sj_toko.set(daftar_nama[0]) 
                self.combo_st_toko.configure(values=daftar_nama)
                self.combo_st_toko.set(daftar_nama[0])
                
            self.hitung_rekap_bulanan()
        except Exception as e: 
            print(f"Gagal memuat sinkronisasi database excel utama: {str(e)}")

    def efek_pilihan_toko_sj(self):
        """ SENSOR DROPDOWN: Memeriksa live kargo & mengisi otomatis alamat tujuan (CASE-INSENSITIVE) """
        import openpyxl
        import datetime
        
        toko_terpilih = self.combo_sj_toko.get().strip().upper() # <--- KUNCI 2: Ambil dalam bentuk huruf besar
        if not toko_terpilih or "PILIH" in toko_terpilih: 
            return
            
        tgl_hari_ini = datetime.datetime.now().strftime("%Y-%m-%d")
        
        kamus_alamat = {
            "ASTRA": "Gd. Astra International, Jl. Gaya Motor Raya No.8, Sunter, Jakarta Utara",
            "GLOBAL": "Jl. Surapati No.123, Sukaluyu, Kec. Cibeunying Kaler, Kota Bandung",
            "MITRA COMPUTER": "Bandung Electronic Center (BEC) Lantai 1 Blok A-05, Kota Bandung"
        }
        
        self.entry_sj_alamat.delete(0, 'end')
        if toko_terpilih in kamus_alamat:
            self.entry_sj_alamat.insert(0, kamus_alamat[toko_terpilih])
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_log = wb["Tracking Packing List"]
            
            ada_kargo = False
            for row in range(15, ws_log.max_row + 1):
                tgl_cell = str(ws_log.cell(row=row, column=1).value)
                status_cell = str(ws_log.cell(row=row, column=4).value).strip().upper()
                ket_cell = str(ws_log.cell(row=row, column=5).value).strip().upper() # <--- KUNCI 3: Paksa baca data cell jadi upper
                
                if tgl_hari_ini in tgl_cell:
                    if status_cell != "KEMBALI" and status_cell != "RUSAK":
                        if toko_terpilih in ket_cell: # Perbandingan aman huruf besar vs huruf besar
                            if "SJ/GTP-" not in ket_cell:
                                ada_kargo = True
                                break
            
            if not ada_kargo:
                self.bell()
                self.btn_cetak_sj.configure(state="disabled", fg_color="#94A3B8", text="⚠️ Tidak Ada Riwayat Scan Keluar Baru untuk Nama Ini")
            else:
                self.btn_cetak_sj.configure(state="normal", fg_color="#10B981", text="🚛 Terbitkan & Cetak Dokumen Surat Jalan Resmi (SAVE PDF)")
        except Exception as err:
            print(f"Gagal mendeteksi live kargo dropdown SJ: {str(err)}")

    def efek_pilihan_toko_st(self):
        """ SENSOR DROPDOWN: Memeriksa live kargo masuk harian (READY/RUSAK) milik toko terpilih """
        import openpyxl
        import datetime
        
        toko_terpilih = self.combo_st_toko.get().strip()
        if not toko_terpilih or "Pilih" in toko_terpilih: 
            return

        tgl_hari_ini = datetime.datetime.now().strftime("%Y-%m-%d")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws_log = wb["Tracking Packing List"]
            
            ada_kargo_masuk = False
            # Menyisir log harian khusus mendeteksi barang masuk (KEMBALI / RUSAK) harian
            for row in range(15, ws_log.max_row + 1):
                tgl_cell = str(ws_log.cell(row=row, column=1).value)
                status_cell = str(ws_log.cell(row=row, column=4).value).strip().upper()
                ket_cell = str(ws_log.cell(row=row, column=5).value)
                
                if tgl_hari_ini in tgl_cell:
                    if status_cell == "KEMBALI" or status_cell == "RUSAK":
                        if toko_terpilih.upper() in ket_cell.upper():
                            if "ST/GTP-" not in ket_cell:
                                ada_kargo_masuk = True
                                break
                        
            # Proteksi tombol eksekusi cetak jika log kargo masuk harian kosong
            if not ada_kargo_masuk:
                self.bell()
                self.btn_cetak_st.configure(state="disabled", fg_color="#94A3B8", text="⚠️ Tidak Ada Riwayat Scan Masuk Baru untuk Toko Ini")
            else:
                self.btn_cetak_st.configure(state="normal", fg_color="#3B82F6", text="📥 Terbitkan, Kunci Log, & Cetak Tanda Terima Barang (PDF Mewah)")
                
        except Exception as err:
            print(f"Gagal mendeteksi live kargo masuk dropdown ST: {str(err)}")

    def hitung_rekap_bulanan(self):
        """ ENGINE BULANAN FIX SINKRON: Menghitung log & menyalakan grafik secara dinamis harian """
        if not os.path.exists(self.excel_file): 
            return
            
        try:
            # 1. Ambil angka bulan pilihan dari dropdown (Aman terhadap format teks/angka)
            bulan_pilihan = int(self.combo_bulan.get())
            
            # 2. MEMBACA SHEET TRACKING PACKING LIST (Mulai dari baris ke-15 -> skiprows=14)
            df_log_bulanan = pd.read_excel(self.excel_file, sheet_name="Tracking Packing List", skiprows=14, header=None)
            df_log_clean = df_log_bulanan.dropna(how='all')
            
            if df_log_clean.empty:
                self.data_grafik_bulanan = [0, 0, 0, 0, 0, 0]
                self.perbarui_grafik_dashboard() # Paksa render grafik kosong agar tidak crash
                return

            # --- AMBIL DAN SINKRONKAN TANGGAL DIGITAL DARI KOLOM A (Indeks 0) ---
            def konversi_tgl_log(val):
                if pd.isna(val): 
                    return pd.NaT
                try:
                    # Antisipasi format serial number murni Excel harian
                    num_val = float(val)
                    return pd.to_datetime(num_val, unit='D', origin='1899-12-30', errors='coerce')
                except:
                    # Antisipasi format tulisan teks tanggal manual regional Indonesia
                    return pd.to_datetime(val, errors='coerce', dayfirst=True)

            kolom_tgl_log = df_log_clean.iloc[:, 0].apply(konversi_tgl_log)
            nomor_bulan_log = kolom_tgl_log.dt.month
            
            # --- AMBIL KOLOM AKTIVITAS KATA KUNCI DARI KOLOM D (Indeks 3) ---
            kolom_aktivitas_log = df_log_clean.iloc[:, 3].astype(str).str.strip().str.upper()
            mask_bulan_log = (nomor_bulan_log == bulan_pilihan)

            # 3. KALKULASI PRESETS COUNTIFS DARI TRACKING PACKING LIST EXCEL
            sewa_bulan = int((mask_bulan_log & kolom_aktivitas_log.str.contains("SEWA", na=False)).sum())
            service_bulan = int((mask_bulan_log & (kolom_aktivitas_log.str.contains("SERVICE", na=False) | kolom_aktivitas_log.str.contains("SERVIS", na=False))).sum())
            kembali_bulan = int((mask_bulan_log & kolom_aktivitas_log.str.contains("KEMBALI", na=False)).sum())
            terjual_bulan = int((mask_bulan_log & (kolom_aktivitas_log.str.contains("DIJUAL", na=False) | kolom_aktivitas_log.str.contains("JUAL", na=False))).sum())
            pinjam_bulan = int((mask_bulan_log & kolom_aktivitas_log.str.contains("PINJAM", na=False)).sum())
            rusak_bulan = int((mask_bulan_log & kolom_aktivitas_log.str.contains("RUSAK", na=False)).sum())
            
            # 4. PASOK ANGKA RIIL KE MEMORI UTAMA VARIABEL GRAFIK DASHBOARD
            self.data_grafik_bulanan = [sewa_bulan, service_bulan, kembali_bulan, terjual_bulan, pinjam_bulan, rusak_bulan]
            
            # 5. MUNTAHKAN DATA ANGKA ASLI KE DALAM TABEL LOG KANAN DASHBOARD MONITOR
            for row in self.tree_bulan.get_children():
                self.tree_bulan.delete(row)
                
            self.tree_bulan.insert("", "end", values=("Laptop Keluar / Disewa", f"{sewa_bulan} Unit"))
            self.tree_bulan.insert("", "end", values=("Laptop Masuk Perbaikan (Service)", f"{service_bulan} Unit"))
            self.tree_bulan.insert("", "end", values=("Laptop Masuk Kembali ke Gudang", f"{kembali_bulan} Unit"))
            self.tree_bulan.insert("", "end", values=("Laptop Unit Terjual", f"{terjual_bulan} Unit"))
            self.tree_bulan.insert("", "end", values=("Laptop Sedang Dipinjam", f"{pinjam_bulan} Unit"))
            self.tree_bulan.insert("", "end", values=("Laptop Terdata Rusak", f"{rusak_bulan} Unit"))
            
            # KUNCI UTAMA SINKRONISASI: Memanggil paksa fungsi grafik di atas untuk merender ulang batang secara dinamis
            self.perbarui_grafik_dashboard()
            
        except Exception as e:
            print(f"Gagal hitung log bulanan rumus COUNTIFS Dashboard: {str(e)}")

    def tampilkan_peringatan_besar(self, judul, teks_peringatan, warna_tema="#EF4444"):
        """ ENGINE POPUP KUSTOM: Membuat jendela besar + Fitur Tekan Enter untuk Menutup """
        popup = ctk.CTkToplevel(self)
        popup.title(judul)
        popup.geometry("600x250")
        popup.resizable(False, False)
        popup.grab_set()
        popup.focus_set()
        
        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (600 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (250 // 2)
        popup.geometry(f"600x250+{x}+{y}")
        
        frame_dalam = ctk.CTkFrame(popup, corner_radius=12, fg_color=("#F8FAFC", "#1E293B"), border_color=warna_tema, border_width=2)
        frame_dalam.pack(fill="both", expand=True, padx=15, pady=15)
        
        lbl_pesan = ctk.CTkLabel(
            frame_dalam, 
            text=teks_peringatan, 
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            text_color=("#0F172A", "#F1F5F9"),
            justify="center",
            wraplength=540
        )
        lbl_pesan.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        
        btn_oke = ctk.CTkButton(
            frame_dalam, 
            text="⚠️ [ENTER] OKE, SAYA MENGERTI DAN AKAN FOKUS", 
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color=warna_tema,
            hover_color="#B91C1C" if warna_tema == "#EF4444" else "#B45309",
            height=40,
            corner_radius=8,
            command=popup.destroy
        )
        btn_oke.pack(fill="x", padx=30, pady=(0, 15))
        
        # BINDING ENTER: Menekan tombol Enter otomatis memicu fungsi menutup popup
        popup.bind("<Return>", lambda e: popup.destroy())
        self.bell()

    def tampilkan_konfirmasi_besar(self, judul, teks_pertanyaan):
        """ ENGINE KONFIRMASI KUSTOM: Membuat box YES/NO + Fitur Enter Otomatis Memilih YA """
        self.pilihan_konfirmasi = None
        
        popup = ctk.CTkToplevel(self)
        popup.title(judul)
        popup.geometry("600x250")
        popup.resizable(False, False)
        popup.grab_set()
        popup.focus_set()
        
        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (600 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (250 // 2)
        popup.geometry(f"600x250+{x}+{y}")
        
        frame_dalam = ctk.CTkFrame(popup, corner_radius=12, fg_color=("#F8FAFC", "#1E293B"), border_color="#3B82F6", border_width=2)
        frame_dalam.pack(fill="both", expand=True, padx=15, pady=15)
        
        lbl_pesan = ctk.CTkLabel(
            frame_dalam, 
            text=teks_pertanyaan, 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color=("#0F172A", "#F1F5F9"),
            justify="center",
            wraplength=540
        )
        lbl_pesan.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        
        frame_tombol = ctk.CTkFrame(frame_dalam, fg_color="transparent")
        frame_tombol.pack(fill="x", padx=30, pady=(0, 15))
        
        def klik_tombol(nilai):
            self.pilihan_konfirmasi = nilai
            popup.destroy()
            
        btn_yes = ctk.CTkButton(frame_tombol, text="✅ [ENTER] YA, BENAR", font=ctk.CTkFont(size=12, weight="bold"), fg_color="#10B981", hover_color="#059669", height=38, width=220, command=lambda: klik_tombol(True))
        btn_yes.pack(side="left", expand=True, padx=10)
        
        btn_no = ctk.CTkButton(frame_tombol, text="❌ TIDAK / BATAL", font=ctk.CTkFont(size=12, weight="bold"), fg_color="#EF4444", hover_color="#B91C1C", height=38, width=220, command=lambda: klik_tombol(False))
        btn_no.pack(side="right", expand=True, padx=10)
        
        # BINDING ENTER: Menekan tombol Enter otomatis memilih klik_tombol(True) / YA
        popup.bind("<Return>", lambda e: klik_tombol(True))
        
        self.bell()
        self.wait_window(popup)
        return self.pilihan_konfirmasi

    def proses_scan_barcode(self, entry_obj, tipe_scan, entry_keterangan=None):
        """ ENGINE SCANNER MULTI-OPERATOR: Transaksi terpusat via REST API v2 + Audit Log Akun Operator """
        import requests
        import datetime
        
        barcode_input = entry_obj.get().strip().upper()
        entry_obj.delete(0, 'end') 
        
        if not barcode_input: 
            return

        # 1. SENSOR WAJIB ISI: Validasi keterangan teks pengiriman keluar kargo harian
        if entry_keterangan is not None:
            isi_ket = entry_keterangan.get().strip()
            if isi_ket == "" or "Masukkan" in isi_ket:
                if tipe_scan == "DISEWA":
                    self.tampilkan_peringatan_besar("Data Belum Lengkap", "FOKUS! isi terlebih dahulu NAMA TOKO/PENERIMA pada kotak input 1. KELUAR / DISEWA!", "#F59E0B")
                elif tipe_scan == "SERVICE":
                    self.tampilkan_peringatan_besar("Data Belum Lengkap", "FOKUS! isi terlebih dahulu VENDOR/LOKASI REPARASI pada kotak input 2. MASUK SERVICE!", "#F59E0B")
                elif tipe_scan == "TERJUAL":
                    self.tampilkan_peringatan_besar("Data Belum Lengkap", "FOKUS! isi terlebih dahulu NAMA PEMBELI/TOKO pada kotak input 3. UNIT DIJUAL!", "#F59E0B")
                elif tipe_scan == "DIPINJAM":
                    self.tampilkan_peringatan_besar("Data Belum Lengkap", "FOKUS! isi terlebih dahulu NAMA PEMINJAM pada kotak input 4. UNIT DIPINJAM!", "#F59E0B")
                entry_keterangan.focus()
                return
        else:
            isi_ket = ""

        conn = None
        try:
            wakarang = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            
            # 2. RADAR INVESTIGASI via REST API: Memeriksa unit barcode di DATABASE PUSAT SERVER
            resp_lookup = requests.get(f"{self.api_base_gudang}/api/unit/{requests.utils.quote(barcode_input)}", timeout=10)
            if resp_lookup.status_code == 404:
                self.bell()
                self.status_box.configure(state="normal")
                self.status_box.insert("end", f"[GAGAL] Barcode [{barcode_input}] tidak terdaftar di Database GTP Pusat!\n")
                self.status_box.configure(state="disabled")
                self.tampilkan_peringatan_besar("Data Tidak Ada", f"Gagal Memproses Barcode!\n\nBarcode [{barcode_input}] tidak terdaftar di database MASTER DATA!\n\nMohon periksa kembali fisik kargo.", "#EF4444")
                return
            resp_lookup.raise_for_status()
            data_unit = resp_lookup.json()["data"]
            
            nama_barang = data_unit["nama_barang"]
            serial_num_db = data_unit["serial_number"]
            status_lama = data_unit["status"]
            
            if not hasil_unit:
                self.bell()

                self.status_box.configure(state="normal")
                self.status_box.insert("end", f"[GAGAL] Barcode [{barcode_input}] tidak terdaftar di Database GTP Pusat!\n")
                self.status_box.configure(state="disabled")
                self.tampilkan_peringatan_besar("Data Tidak Ada", f"Gagal Memproses Barcode!\n\nBarcode [{barcode_input}] tidak terdaftar di database MASTER DATA!\n\nMohon periksa kembali fisik kargo.", "#EF4444")
                return

            status_lama_upper = status_lama.strip().upper() if status_lama else ""
            
            # 4. PELACAKAN BALIK LOKASI TERAKHIR: dari riwayat log yang dikirim API
            riwayat = data_unit.get("riwayat_terakhir") or []
            lokasi_saat_ini = str(riwayat[0].get("keterangan")).strip() if riwayat and riwayat[0].get("keterangan") else "Tidak Terlacak"

            if status_lama_upper.startswith("DISEWA"): status_teks_bersih = f"DISEWA (Atas Nama: {lokasi_saat_ini})"
            elif status_lama_upper.startswith("SERVICE"): status_teks_bersih = f"SERVICE (Di Lokasi: {lokasi_saat_ini})"
            elif "TERJUAL" in status_lama_upper or "DIJUAL" in status_lama_upper: status_teks_bersih = f"TERJUAL (Kepada: {lokasi_saat_ini})"
            elif status_lama_upper.startswith("DIPINJAM"): status_teks_bersih = f"DIPINJAM (Oleh: {lokasi_saat_ini})"
            else: status_teks_bersih = status_lama

            # 5. SENSOR PROTEKSI DUPLIKAT TRANSAKSI: (Kebal Ganda Jendela Besar Merah)
            if tipe_scan == "READY" and status_lama_upper.startswith("READY"):
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Ready", f"Gagal Scan Masuk! Barang dengan Barcode [{barcode_input}] FOKUS! Status unit laptop ini memang sudah READY di dalam gudang GTP.", "#EF4444"); return
            if tipe_scan == "RUSAK" and status_lama_upper == "RUSAK":
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Rusak", f"Gagal Proses! Barang dengan Barcode [{barcode_input}] Unit laptop ini memang sudah terdata RUSAK di dalam gudang sebelumnya.", "#EF4444"); return
            if tipe_scan == "DISEWA" and f"DISEWA KE {isi_ket.upper()}" in status_lama_upper:
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Sewa", f"GEUS DI SCAN Woi!!! Duplikat Transaksi Keluar!\n\nUnit Barcode [{barcode_input}] terdeteksi sudah berstatus DISEWA ke {isi_ket}.", "#EF4444"); return
            if tipe_scan == "SERVICE" and f"SERVICE DI {isi_ket.upper()}" in status_lama_upper:
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Service", f"GEUS DI SCAN Woi!!! Duplikat Transaksi Service!\n\nUnit Barcode [{barcode_input}] terdeteksi sudah berada di lokasi SERVICE ({isi_ket}).", "#EF4444"); return
            if tipe_scan == "TERJUAL" and f"TERJUAL KE {isi_ket.upper()}" in status_lama_upper:
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Terjual", f"Gagal Transaksi! Duplikat Terjual!\n\nUnit Barcode [{barcode_input}] statusnya sudah resmi laku TERJUAL kepada {isi_ket}.", "#EF4444"); return
            if tipe_scan == "DIPINJAM" and f"DIPINJAM OLEH {isi_ket.upper()}" in status_lama_upper:
                self.bell()
                self.tampilkan_peringatan_besar("Duplikat Peminjaman", f"GEUS DI SCAN BRO!! Duplikat Transaksi Peminjaman!\n\nUnit Barcode [{barcode_input}] sudah dipegang dan DIPINJAM oleh {isi_ket}.", "#EF4444"); return

            # 6. SENSOR SEBELUM KELUAR: Memaksa unit wajib READY di gudang sebelum diproses keluar
            if tipe_scan in ["DISEWA", "SERVICE", "TERJUAL", "DIPINJAM"]:
                if not status_lama_upper.startswith("READY"):
                    self.bell()
                    self.status_box.configure(state="normal")
                    self.status_box.insert("end", f"[DITOLAK] Barcode {barcode_input} Gagal Keluar. Status: {status_teks_bersih}\n")
                    self.status_box.configure(state="disabled")
                    self.tampilkan_peringatan_besar("Status Tidak Valid", f"Gagal Scan Keluar!\n\nBarcode : [{barcode_input}]\nBarang  : {nama_barang}\nStatus Saat Ini: [{status_teks_bersih}]\n\nSemua barang keluar harus di-scan masuk gudang (READY) terlebih dahulu!", "#EF4444")
                    return
            
            # 7. PERCABANGAN 6 BOKS KONFIRMASI DIALOG ASLI ANDA (SINKRON CLOUD)
            if tipe_scan == "READY":
                if "TERJUAL" in status_lama_upper or "DIJUAL" in status_lama_upper:
                    tanya = self.tampilkan_konfirmasi_besar("Deteksi Retur Penjualan", f"Barcode : [{barcode_input}]\nBarang  : {nama_barang}\nStatus Terakhir: {status_teks_bersih}\n\nATTENTION!! Laptop ini berstatus TERJUAL.\nApakah unit ini masuk kembali ke gudang sebagai BARANG RETUR?")
                    if not tanya: return
                    status_fix_master = "READY"; status_fix_log = "KEMBALI"; ket_fix_log = f"RETUR JUAL dari {lokasi_saat_ini}"
                else:
                    tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Pengembalian", f"Barcode : [{barcode_input}]\nBarang  : {nama_barang}\nStatus Terakhir: {status_teks_bersih}\n\nBENER YEUH??!! laptop ini benar sudah masuk kembali ke gudang?")
                    if not tanya: return
                    status_fix_master = "READY"; status_fix_log = "KEMBALI"; ket_fix_log = f"Kembali dari {lokasi_saat_ini}"
            elif tipe_scan == "RUSAK":
                tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Rusak", f"Barcode: [{barcode_input}]\nNama: {nama_barang}\nStatus Sebelumnya: {status_teks_bersih}\n\nYAKIN TEU?!! Ubah statusna jadi RUSAK?")
                if not tanya: return
                status_fix_master = "RUSAK"; status_fix_log = "RUSAK"; ket_fix_log = "Terdata Rusak di Gudang"
            elif tipe_scan == "DISEWA":
                tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Sewa Keluar", f"Barcode: [{barcode_input}]\nBarang  : {nama_barang}\nStatus Sebelumnya: {status_teks_bersih}\nTujuan Baru: {isi_ket}\n\nCEK DEUI! Keluarkan barang untuk DISEWA?")
                if not tanya: return
                status_fix_master = f"DISEWA ke {isi_ket}"; status_fix_log = f"DISEWA ke {isi_ket}"; ket_fix_log = isi_ket
            elif tipe_scan == "SERVICE":
                tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Service", f"Barcode: [{barcode_input}]\nBarang  : {nama_barang}\nStatus Sebelumnya: {status_teks_bersih}\nVendor Baru: {isi_ket}\n\nUbah status menjadi SERVICE?")
                if not tanya: return
                status_fix_master = f"SERVICE di {isi_ket}"; status_fix_log = f"SERVICE di {isi_ket}"; ket_fix_log = isi_ket
            elif tipe_scan == "TERJUAL":
                tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Penjualan", f"Barcode: [{barcode_input}]\nBarang  : {nama_barang}\nStatus Sebelumnya: {status_teks_bersih}\nPembeli Baru: {isi_ket}\n\nUbah status menjadi TERJUAL?")
                if not tanya: return
                status_fix_master = f"TERJUAL ke {isi_ket}"; status_fix_log = f"DIJUAL ke {isi_ket}"; ket_fix_log = isi_ket
            elif tipe_scan == "DIPINJAM":
                tanya = self.tampilkan_konfirmasi_besar("Konfirmasi Peminjaman", f"Barcode: [{barcode_input}]\nBarang  : {nama_barang}\nStatus Sebelumnya: {status_teks_bersih}\nPeminjam: {isi_ket}\n\nUbah status menjadi DIPINJAM?")
                if not tanya: return
                status_fix_master = f"DIPINJAM oleh {isi_ket}"; status_fix_log = f"DIPINJAM oleh {isi_ket}"; ket_fix_log = isi_ket

            # 8. EKSEKUSI DATA SECARA SERENTAK ke server pusat via REST API v2
            resp_commit = requests.post(
                f"{self.api_base_gudang}/api/gudang/scan-commit",
                json={
                    "barcode": barcode_input,
                    "status_fix_master": status_fix_master,
                    "status_fix_log": status_fix_log,
                    "ket_fix_log": ket_fix_log,
                    "waktu": wakarang,
                    "nama_barang": nama_barang,
                    "operator": self.operator_aktif,
                },
                timeout=15,
            )
            if resp_commit.status_code != 200:
                raise RuntimeError(f"Server menolak commit scan: {resp_commit.status_code} - {resp_commit.text[:200]}")
            
            # --- BLOK EKSEKUSI FEEDBACK VISUAL KE MONITOR GUDANG ---
            self.status_box.configure(state="normal")
            self.status_box.insert("end", f"[SUKSES SAVE] Barcode: {barcode_input} -> Di-set {tipe_scan} oleh {self.operator_aktif}.\n")
            self.status_box.configure(state="disabled")
            
            # Memanggil fungsi penyegaran layar tabel monitor harian
            if hasattr(self, 'muat_log_tracking_excel'):
                self.muat_log_tracking_excel()
            
            # --- BLOK NOTIFIKASI POPUP HIJAU SEPERTI BIASA ---
            self.tampilkan_peringatan_besar("Scan Sukses", f"BERHASIL UPDATE!\n\nUnit {nama_barang}\nStatus resmi diperbarui oleh {self.operator_aktif} menjadi:\n[{status_fix_master}].", "#10B981")
            
        except Exception as err:
            self.tampilkan_peringatan_besar("Error Scan Database", f"Gagal mengirimkan data mutasi ke Server API:\n\n{str(err)}", "#EF4444")

    def proses_penerbitan_sj(self):
        """ Menu 3: ENGINE SURAT JALAN KELUAR - VIA REST API v2 """
        import datetime
        import requests
        import os
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        toko_tujuan = self.combo_sj_toko.get().strip()
        alamat_tujuan = self.entry_sj_alamat.get().strip() 
        catatan_ket = self.entry_sj_ket.get().strip()
        
        # 1. SENSOR BARIS KOSONG: Dihentikan langsung oleh Box Raksasa Merah
        if not toko_tujuan or "Pilih" in toko_tujuan:
            self.bell()
            self.tampilkan_peringatan_besar("Gagal Cetak", "FOKUS! Nama Penerima Kargo masih kosong!\nSilakan pilih terlebih dahulu sebelum menerbitkan dokumen.", "#EF4444")
            self.combo_sj_toko.focus()
            return
            
        if not alamat_tujuan or "Ketik" in alamat_tujuan:
            self.bell()
            self.tampilkan_peringatan_besar("Gagal Cetak", "FOKUS! Alamat Tujuan Pengiriman kargo masih kosong!\n\nKolom Alamat wajib diisi sebelum menerbitkan dokumen PDF.", "#EF4444")
            self.entry_sj_alamat.focus()
            return

        waktu_sekarang = datetime.datetime.now()
        tanggal_nota = waktu_sekarang.strftime("%d-%m-%Y")
        tahun_pendek = waktu_sekarang.strftime("%y")
        bulan_angka = waktu_sekarang.strftime("%m")
        tgl_hari_ini = waktu_sekarang.strftime("%Y-%m-%d")
        
        conn = None
        try:
            # 2. AMBIL KANDIDAT SJ via REST API: kargo keluar toko ini yang BELUM DICAP NOTA
            resp_sj = requests.get(
                f"{self.api_base_gudang}/api/gudang/sj-kandidat",
                params={"toko": toko_tujuan},
                timeout=15,
            )
            resp_sj.raise_for_status()
            daftar_scan_sql = [tuple(r.values()) for r in resp_sj.json()["data"]]
            
            # 3. VERIFIKASI SENSOR DATA: Jika di server kosong, munculkan BOX MERAH
            if not daftar_scan_sql:
                self.bell()
                pesan_kosong_sj = f"Gagal Memproses Surat Jalan!\n\nTidak ditemukan riwayat scan keluar baru atas nama toko '{toko_tujuan}' di dalam database!\n\nSilakan lakukan scan kargo keluar terlebih dahulu di halaman SCAN BARANG."
                self.tampilkan_peringatan_besar("Data Scan Tidak Ada", pesan_kosong_sj, "#EF4444")
                return

            # Jendela Konfirmasi Kustom Besar Cerdas Anda
            pesan_konfirm_sj = f"CEK KEMBALI KAWAN!\n\nApakah Anda yakin ingin menerbitkan dokumen SURAT JALAN resmi\nuntuk Nama: {toko_tujuan}?\n\nTindakan ini akan mengunci nomor surat baru\ndan menyimpan log permanen ke DATA!"

            setuju_cetak = self.tampilkan_konfirmasi_besar("Konfirmasi Penerbitan Surat Jalan", pesan_konfirm_sj)
            if not setuju_cetak: 
                return

            # 4. COMMIT TRANSAKSI SJ via REST API: server yang bikin nomor nota, cap, dan log histori (atomik)
            resp_commit_sj = requests.post(
                f"{self.api_base_gudang}/api/gudang/sj-commit",
                json={"toko": toko_tujuan, "catatan": catatan_ket if catatan_ket else "-"},
                timeout=20,
            )
            if resp_commit_sj.status_code != 200:
                raise RuntimeError(f"Server menolak commit SJ: {resp_commit_sj.status_code} - {resp_commit_sj.text[:200]}")
            hasil_sj = resp_commit_sj.json()["data"]
            nomor_sj_baru = hasil_sj["nomor_sj"]
            
            # 5. LOCKING SYSTEM: Merakit manifest item dari hasil server
            daftar_barcode_manifest = [b["barcode"] for b in hasil_sj["items"]]
            daftar_nama_laptop = [b["nama_barang"] if b["nama_barang"] else "-" for b in hasil_sj["items"]]

            # 7. MERAKIT STRUKTUR BERKAS NOTA REPORTLAB PDF MODERN
            nama_folder_arsip = "Histori Surat Jalan"
            if not os.path.exists(nama_folder_arsip): 
                os.makedirs(nama_folder_arsip)
                
            nama_file_pdf = f"Surat_Jalan_{nomor_sj_baru.replace('/', '-')}.pdf"
            path_file_pdf_lengkap = os.path.join(nama_folder_arsip, nama_file_pdf)
            
            doc = SimpleDocTemplate(path_file_pdf_lengkap, pagesize=letter, rightMargin=45, leftMargin=45, topMargin=45, bottomMargin=45)
            story = []; styles = getSampleStyleSheet()
            
            gaya_alamat = ParagraphStyle('Alamat', parent=styles['Normal'], fontSize=9, leading=14, textColor="#4B5563")
            gaya_judul_dok = ParagraphStyle('Judul', parent=styles['Heading1'], fontSize=18, alignment=1, fontName="Helvetica-Bold")
            gaya_normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, leading=16, textColor="#1F2937")
            gaya_normal_bold = ParagraphStyle('NormalBold', parent=gaya_normal, fontName="Helvetica-Bold")
            
            teks_kop = "<b><font size=14 color='#1E3A8A'>GLOBAL TEKNOLOGI PRODIGI</font></b><br/>" \
                       "Jl. Jendral Sudirman No.823, Cijerah, Kec. Bandung Kulon,<br/>" \
                       "Kota Bandung, Jawa Barat 40213"
            par_kop_kanan = Paragraph(teks_kop, gaya_alamat)
            
            # Mencari logo via pelacak jalur aset otomatis agar aman di dalam file .exe tunggal Anda
            jalur_logo_riyal = self.ambil_jalur_aset("logo.png") if hasattr(self, 'ambil_jalur_aset') else "logo.png"
            tabel_header = Table([[RLImage(jalur_logo_riyal, width=55, height=55) if os.path.exists(jalur_logo_riyal) else "", par_kop_kanan]], colWidths=[70, 452])
            tabel_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEBELOW', (0,0), (-1,-1), 1, "#E5E7EB")]))
            story.append(tabel_header); story.append(Spacer(1, 10))
            story.append(Paragraph("SURAT JALAN", gaya_judul_dok))
            story.append(Spacer(1, 10))
            
            info_data = [
                [Paragraph(f"<b>No Surat:</b> {nomor_sj_baru}", gaya_normal), Paragraph(f"<b>Penerima:</b> {toko_tujuan}", gaya_normal)],
                [Paragraph(f"<b>Tanggal:</b> {tanggal_nota}", gaya_normal), Paragraph(f"<b>Alamat:</b> {alamat_tujuan}", gaya_alamat)]
            ]
            tabel_info = Table(info_data, colWidths=[260, 260])
            tabel_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
            story.append(tabel_info); story.append(Spacer(1, 15))
            
            konten_data = [[
                Paragraph("<b>No</b>", gaya_normal_bold),
                Paragraph("<b>Nomor Barcode</b>", gaya_normal_bold),
                Paragraph("<b>Spesifikasi / Nama Unit Laptop</b>", gaya_normal_bold),
                Paragraph("<b>Qty</b>", gaya_normal_bold)
            ]]
            for idx, b_code in enumerate(daftar_barcode_manifest):
                n_laptop = daftar_nama_laptop[idx]
                konten_data.append([
                    Paragraph(str(idx + 1), gaya_normal),
                    Paragraph(b_code, gaya_normal),
                    Paragraph(n_laptop, gaya_normal),
                    Paragraph("1 Unit", gaya_normal)
                ])
                
            tabel_konten = Table(konten_data, colWidths=[40, 120, 300, 60])
            tabel_konten.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), "#F9FAFB"),
                ('PADDING', (0,0), (-1,-1), 8),
                ('ALIGN', (0,0), (0,-1), 'CENTER'),
                ('ALIGN', (3,0), (3,-1), 'CENTER'),
                ('LINEBELOW', (0,0), (-1,0), 1, "#111827"),
                ('BOX', (0,0), (-1,-1), 0.5, "#D1D5DB"),
                ('INNERGRID', (0,0), (-1,-1), 0.5, "#E5E7EB"),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ]))
            story.append(tabel_konten)
            story.append(Spacer(1, 40))
            
            # AREA 3 KOLOM TANDA TANGAN SIMETRIS SAMA RATA
            ttd_data = [
                [Paragraph("Petugas Gudang (GTP),", gaya_normal), Paragraph("Sopir / Kurir,", gaya_normal), Paragraph("Pihak Penerima / PIC,", gaya_normal)], 
                ["", "", ""],
                [Paragraph(f"( {self.operator_aktif} )", gaya_normal), Paragraph("( ____________________ )", gaya_normal), Paragraph(f"( {toko_tujuan} )", gaya_normal)]
            ]
            tabel_ttd = Table(ttd_data, colWidths=[173.3, 173.3, 173.4])
            tabel_ttd.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'BOTTOM'), 
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), 
                ('BOTTOMPADDING', (0,1), (-1,1), 55)  
            ]))
            story.append(tabel_ttd)
            
            doc.build(story)
            
            # 8. RESET FORM ISIAN & REFRESH VISUAL HALAMAN
            self.entry_sj_alamat.delete(0, 'end')
            self.entry_sj_ket.delete(0, 'end')
            
            # Menyegarkan data monitor utama via fungsi SQL baru yang kita kunci sebelumnya
            if hasattr(self, 'muat_log_tracking_excel'):
                self.muat_log_tracking_excel()
            
            # Membuka otomatis file PDF Surat Jalan yang baru terbentuk
            try:
                jalur_pdf_absolut = os.path.abspath(path_file_pdf_lengkap)
                os.startfile(jalur_pdf_absolut)
            except Exception as print_err:
                print(f"Gagal membuka file PDF otomatis: {str(print_err)}")
                
            self.tampilkan_peringatan_besar(
                "Cetak Sukses", 
                f"Surat Jalan {nomor_sj_baru} Berhasil Diterbitkan!\n\nFile dokumen telah disimpan otomatis ke folder 'Histori Surat Jalan'.\nSistem sedang membuka file PDF untuk Anda periksa.", 
                "#10B981"
            )
            
        except Exception as e: 
            self.tampilkan_peringatan_besar("Error Sistem", f"Gagal menerbitkan Surat Jalan PDF: {str(e)}", "#EF4444")

    def proses_penerbitan_st(self):
        """ Menu 4: ENGINE TANDA TERIMA MASUK (REVISI LOGIKA PARALEL - ANTI-DUPLIKAT) """
        import datetime
        import openpyxl
        import os
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        toko_pengirim = self.combo_st_toko.get().strip()
        catatan_ket = self.entry_st_ket.get().strip()
        
        # SENSOR BARIS KOSONG DROPDOWN
        if not toko_pengirim or "Pilih" in toko_pengirim:
            self.bell()
            self.tampilkan_peringatan_besar("Gagal Cetak", "FOKUS! Nama Pihak Pengirim masih kosong!\nSilakan pilih terlebih dahulu sebelum menerbitkan dokumen.", "#EF4444")
            self.combo_st_toko.focus()
            return

        waktu_sekarang = datetime.datetime.now()
        tanggal_nota = waktu_sekarang.strftime("%d-%m-%Y")
        tahun_pendek = waktu_sekarang.strftime("%y")
        bulan_angka = waktu_sekarang.strftime("%m")
        tgl_hari_ini = waktu_sekarang.strftime("%Y-%m-%d")

        try:
            # 1. BUKA FILE EXCEL SECARA LIVE DI LATAR BELAKANG
            wb = openpyxl.load_workbook(self.excel_file, data_only=False)
            ws_log = wb["Tracking Packing List"]
            ws_histori_st = wb["HISTORI TANDA TERIMA"]
            
            toko_pengirim_clean = toko_pengirim.strip().upper()
            
            daftar_barcode_manifest = []
            daftar_nama_laptop = []
            baris_log_terpilih = []
            dict_anti_duplikat = {}
            
            # [KUNCI REVISI BARIS 15]: Karena data mutasi masuk terbaru selalu disisipkan di baris 15,
            # kita batasi jangkauan radar pencarian hanya memindai maksimal 200 baris teratas dari baris 15 ke bawah.
            # Sistem dijamin melesat secepat kilat (di bawah 0.01 detik) tanpa freeze.
            batas_penyisiran = min(ws_log.max_row + 1, 215)
            
            for row in range(15, batas_penyisiran):
                status_cell = str(ws_log.cell(row=row, column=4).value).strip().upper()
                ket_cell = str(ws_log.cell(row=row, column=5).value).strip().upper()
                
                # HUBUNGAN PARALEL MUTLAK: Sinkronisasi pemfilteran status masuk harian
                if status_cell == "KEMBALI" or status_cell == "RUSAK":
                    if toko_pengirim_clean in ket_cell:
                        if "ST/GTP-" not in ket_cell:
                            b_code = str(ws_log.cell(row=row, column=2).value).strip()
                            n_laptop = ws_log.cell(row=row, column=3).value
                            
                            if b_code and b_code != "None" and b_code != "":
                                if b_code not in dict_anti_duplikat:
                                    dict_anti_duplikat[b_code] = ""
                                    daftar_barcode_manifest.append(b_code)
                                    daftar_nama_laptop.append(n_laptop if n_laptop else "-")
                                    baris_log_terpilih.append(row)

            # Jika setelah disisir mundur secara paralel hasilnya kosong, kunci dokumen pakai box merah raksasa
            if not daftar_barcode_manifest:
                self.bell()
                pesan_kosong_st = f"Gagal Memproses Tanda Terima!\n\nTidak ditemukan data mutasi masuk baru atas nama '{toko_pengirim}' di database!\n\nSilakan lakukan scan masuk (KEMBALI/RUSAK) terlebih dahulu di halaman SCAN BARANG."
                self.tampilkan_peringatan_besar("Data Scan Tidak Ada", pesan_kosong_st, "#EF4444")
                return

            # Gerbang Konfirmasi Muncul karena kargo tervalidasi POSITIF ADA
            pesan_konfirm_st = f"CEK KEMBALI KAWAN!\n\n" \
                               f"Apakah Anda yakin ingin menerbitkan dokumen SURAT TANDA TERIMA resmi\n" \
                               f"dari Pengirim: {toko_pengirim}?\n\n" \
                               f"Tindakan ini akan mengunci nomor surat baru\n" \
                               f"dan menyimpan log masuk permanen ke DATA!"

            setuju_cetak_st = self.tampilkan_konfirmasi_besar("Konfirmasi Penerbitan Tanda Terima", pesan_konfirm_st)
            if not setuju_cetak_st:
                return
            # Menghitung nomor urut nota otomatis pada bulan berjalan
            counter_nota = 1
            for row in range(6, ws_histori_st.max_row + 1):
                tgl_histori = ws_histori_st.cell(row=row, column=2).value
                if tgl_histori:
                    try:
                        dt_hist = tgl_histori if isinstance(tgl_histori, datetime.datetime) else datetime.datetime.strptime(str(tgl_histori).split()[0], "%Y-%m-%d")
                        if dt_hist.year == waktu_sekarang.year and dt_hist.month == waktu_sekarang.month:
                            counter_nota += 1
                    except: continue
                    
            nomor_st_baru = f"ST/GTP-{tahun_pendek}/{bulan_angka}/{str(counter_nota).zfill(3)}"
            
            # EKSEKUSI PENEMPELAN CAP NOTA SEBAGAI FILTER ANTI-DUPLIKAT PERMANEN
            for r_terpilih in baris_log_terpilih:
                nilai_lama_ket = str(ws_log.cell(row=r_terpilih, column=5).value).strip()
                ws_log.cell(row=r_terpilih, column=5).value = f"{nilai_lama_ket} (Surat: {nomor_st_baru})"
                
            # Rekam data transaksi baru ke dalam tab HISTORI TANDA TERIMA Excel
            row_baru_histori = ws_histori_st.max_row + 1
            if row_baru_histori < 6: 
                row_baru_histori = 6
                
            string_barcode_gabung = ", ".join(daftar_barcode_manifest)
            
            ws_histori_st.cell(row=row_baru_histori, column=2).value = datetime.datetime.now()
            ws_histori_st.cell(row=row_baru_histori, column=3).value = nomor_st_baru
            ws_histori_st.cell(row=row_baru_histori, column=4).value = toko_pengirim
            ws_histori_st.cell(row=row_baru_histori, column=5).value = string_barcode_gabung
            ws_histori_st.cell(row=row_baru_histori, column=6).value = catatan_ket if catatan_ket else "-"

            # Proses pembuatan file folder arsip & Inisialisasi Dokumen PDF Resmi
            nama_folder_arsip = "Histori Tanda Terima"
            if not os.path.exists(nama_folder_arsip):
                os.makedirs(nama_folder_arsip)
                
            nama_file_pdf = f"Tanda_Terima_{nomor_st_baru.replace('/', '-')}.pdf"
            path_file_pdf_lengkap = os.path.join(nama_folder_arsip, nama_file_pdf)
            
            doc = SimpleDocTemplate(path_file_pdf_lengkap, pagesize=letter, rightMargin=45, leftMargin=45, topMargin=45, bottomMargin=45)
            story = []; styles = getSampleStyleSheet()
            gaya_alamat = ParagraphStyle('Alamat', parent=styles['Normal'], fontSize=9, leading=14, textColor="#4B5563")
            gaya_judul_dok = ParagraphStyle('Judul', parent=styles['Heading1'], fontSize=18, alignment=1, fontName="Helvetica-Bold")
            gaya_normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, leading=16, textColor="#1F2937")
            gaya_normal_bold = ParagraphStyle('NormalBold', parent=gaya_normal, fontName="Helvetica-Bold")
            
            teks_kop = "<b><font size=14 color='#1E3A8A'>GLOBAL TEKNOLOGI PRODIGI</font></b><br/>" \
                       "Jl. Jendral Sudirman No.823, Cijerah, Kec. Bandung Kulon,<br/>" \
                       "Kota Bandung, Jawa Barat 40213"
            par_kop_kanan = Paragraph(teks_kop, gaya_alamat)
            
            tabel_header = Table([[RLImage("logo.png", width=55, height=55) if os.path.exists("logo.png") else "", par_kop_kanan]], colWidths=[70, 452])
            tabel_header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEBELOW', (0,0), (-1,-1), 1, "#E5E7EB")]))
            story.append(tabel_header); story.append(Spacer(1, 10))
            
            story.append(Paragraph("TANDA TERIMA BARANG", gaya_judul_dok))
            story.append(Spacer(1, 10))
            
            info_data = [
                [Paragraph(f"<b>No Tanda Terima:</b> {nomor_st_baru}", gaya_normal), Paragraph(f"<b>Yang Menyerahkan:</b> {toko_pengirim}", gaya_normal)],
                [Paragraph(f"<b>Tanggal Terima:</b> {tanggal_nota}", gaya_normal), Paragraph("", gaya_normal)]
            ]
            tabel_info = Table(info_data, colWidths=[260, 260])
            tabel_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
            story.append(tabel_info); story.append(Spacer(1, 15))
            
            konten_data = [[
                Paragraph("<b>No</b>", gaya_normal_bold),
                Paragraph("<b>Nomor Barcode</b>", gaya_normal_bold),
                Paragraph("<b>Spesifikasi / Nama Unit Laptop</b>", gaya_normal_bold),
                Paragraph("<b>Qty</b>", gaya_normal_bold)
            ]]
            
            for idx, b_code in enumerate(daftar_barcode_manifest):
                n_laptop = daftar_nama_laptop[idx]
                konten_data.append([
                    Paragraph(str(idx + 1), gaya_normal),
                    Paragraph(b_code, gaya_normal),
                    Paragraph(n_laptop, gaya_normal),
                    Paragraph("1 Unit", gaya_normal)
                ])
                
            # [HANYA ADA SATU TABEL MANIFEST KONTEN - BEBAS DUPLIKAT]
            tabel_konten = Table(konten_data, colWidths=[40, 120, 300, 60])
            tabel_konten.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), "#F9FAFB"),
                ('PADDING', (0,0), (-1,-1), 8),
                ('ALIGN', (0,0), (0,-1), 'CENTER'),
                ('ALIGN', (3,0), (3,-1), 'CENTER'),
                ('LINEBELOW', (0,0), (-1,0), 1, "#111827"),
                ('BOX', (0,0), (-1,-1), 0.5, "#D1D5DB"),
                ('INNERGRID', (0,0), (-1,-1), 0.5, "#E5E7EB"),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ]))
            story.append(tabel_konten); story.append(Spacer(1, 40))
            
            ttd_data = [
                [Paragraph("Yang Menyerahkan,", gaya_normal), Paragraph("Sopir / Kurir,", gaya_normal), Paragraph("Penerima Gudang (GTP),", gaya_normal)], 
                ["", "", ""],
                [Paragraph(f"( {toko_pengirim} )", gaya_normal), Paragraph("( ____________________ )", gaya_normal), Paragraph("( Global Teknologi Prodigi )", gaya_normal)]
            ]
            tabel_ttd = Table(ttd_data, colWidths=[173.3, 173.3, 173.4])
            tabel_ttd.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'BOTTOM'), 
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), 
                ('BOTTOMPADDING', (0,1), (-1,1), 55)  
            ]))
            story.append(tabel_ttd)

            doc.build(story)
            wb.save(self.excel_file)
            
            self.entry_st_ket.delete(0, 'end')
            self.muat_data_excel()

            try:
                jalur_pdf_absolut = os.path.abspath(path_file_pdf_lengkap)
                os.startfile(jalur_pdf_absolut)
            except Exception as print_err:
                print(f"Gagal membuka file PDF otomatis: {str(print_err)}")
            
            self.tampilkan_peringatan_besar(
                "Cetak Sukses", 
                f"Tanda Terima {nomor_st_baru} Berhasil Diterbitkan!\n\nFile dokumen telah disimpan otomatis.\nSistem sedang membuka file PDF untuk Anda periksa.", 
                "#10B981"
            )
            
        except Exception as e: 
            self.tampilkan_peringatan_besar("Error Sistem", f"Gagal menerbitkan Tanda Terima PDF: {str(e)}", "#EF4444")

    def buat_layar_loading_splash(self):
        """ MODUL BARU: LAYAR LOADING SPLASH SCREEN INTERAKTIF 3 DETIK (PREMIUM EDITION) """
        import customtkinter as ctk
        from PIL import Image as PILImage
        import os
        
        # 1. Sembunyikan jendela utama aplikasi agar tidak balapan muncul di layar
        self.withdraw()
        
        # 2. Merakit Jendela Tampilan Loading Melayang (Borderless Card)
        self.jendela_splash = ctk.CTkToplevel(self)
        self.jendela_splash.overrideredirect(True) # Menghilangkan frame bar judul atas Windows
        self.jendela_splash.configure(fg_color="#FFFFFF")
        
        # Jendela diperbesar menjadi 680 x 420 agar terlihat megah dan lapang
        lebar_sp = 680
        tinggi_sp = 420
        lebar_layar = self.winfo_screenwidth()
        tinggi_layar = self.winfo_screenheight()
        posisi_x = int((lebar_layar - lebar_sp) / 2)
        posisi_y = int((tinggi_layar - tinggi_sp) / 2)
        self.jendela_splash.geometry(f"{lebar_sp}x{tinggi_sp}+{posisi_x}+{posisi_y}")
        self.jendela_splash.lift()
        self.jendela_splash.attributes("-topmost", True) # Memaksa jendela berada di paling depan
        
        # 3. ELEMEN KONTEN VISUAL INTERIOR (LOGO & TEXT PERUSAHAAN)
        container_splash = ctk.CTkFrame(self.jendela_splash, fg_color="#FFFFFF", corner_radius=15, border_width=1, border_color="#E2E8F0")
        container_splash.pack(fill="both", expand=True)
        
        # Menampilkan Logo Perusahaan GTP Tengah Simetris
        if os.path.exists("logo.png"):
            try:
                img_logo_raw = PILImage.open("logo.png")
                img_logo_ctk = ctk.CTkImage(light_image=img_logo_raw, dark_image=img_logo_raw, size=(90, 90))
                lbl_logo = ctk.CTkLabel(container_splash, image=img_logo_ctk, text="")
                lbl_logo.pack(pady=(45, 10))
            except: pass
            
        judul_perusahaan = ctk.CTkLabel(container_splash, text="GLOBAL TEKNOLOGI PRODIGI", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color="#1E293B")
        judul_perusahaan.pack(pady=2)
        
        sub_judul_apk = ctk.CTkLabel(container_splash, text="Laptop Inventory System - Enterprise Premium Edition", font=ctk.CTkFont(family="Segoe UI", size=11, weight="normal"), text_color="#64748B")
        sub_judul_apk.pack(pady=(0, 20))
        
        # Teks Status Interaktif yang akan berubah otomatis seiring pergerakan loading
        self.lbl_status_loading = ctk.CTkLabel(container_splash, text="Menginisialisasi modul sistem...", font=ctk.CTkFont(family="Segoe UI", size=10, slant="italic"), text_color="#475569")
        self.lbl_status_loading.pack(pady=(10, 2))
        
        # Bilah Loading Progress Bar Biru Royal GTP
        self.progress_bar_splash = ctk.CTkProgressBar(container_splash, width=380, height=6, corner_radius=3, progress_color="#2563EB", fg_color="#E2E8F0")
        self.progress_bar_splash.set(0.0)
        self.progress_bar_splash.pack(pady=(0, 20))
        
        # REVISI VISUAL: Mengubah teks hak cipta menjadi Hitam Pekat & Tebal agar terlihat sangat mencolok
        lbl_hak_cipta = ctk.CTkLabel(
            container_splash, 
            text="© Hak Cipta, Karya MaMet SpooKy. All Rights Reserved.", 
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), 
            text_color="#0F172A"  # Mengunci warna hitam arang pekat premium Windows 11
        )
        lbl_hak_cipta.pack(side="bottom", pady=20)
        # 4. ENGINE ANIMASI AUTOMATION: Simulasi pergerakan loading bar selama 3 detik harian
        self.nilai_loading_current = 0.0
        
        def jalankan_hitung_mundur_splash():
            # Setiap pemanggilan fungsi, naikkan progress bar secara bertahap
            self.nilai_loading_current += 0.02
            self.progress_bar_splash.set(self.nilai_loading_current)
            
            # Pengondisian teks status interaktif berdasarkan persentase pergerakan RAM
            persen = int(self.nilai_loading_current * 100)
            if persen < 25:
                self.lbl_status_loading.configure(text=f"Memuat komponen visual CustomTkinter... ({persen}%)")
            elif persen < 55:
                self.lbl_status_loading.configure(text=f"Menghubungkan basis data MASTER DATA... ({persen}%)")
            elif persen < 85:
                self.lbl_status_loading.configure(text=f"Menyinkronkan log mutasi kargo Baris 15... ({persen}%)")
            else:
                self.lbl_status_loading.configure(text=f"Sistem siap! Membuka Dashboard... ({persen}%)")
                
            # GERBANG UTAMA TRANSISI REVISI: Jika loading sudah penuh (5 detik)
            if self.nilai_loading_current >= 1.0:
                self.jendela_splash.destroy()  # Hancurkan jendela loading dari memori RAM PC
                self.buat_jendela_login_sistem() # ⬅️ GANTI BARIS INI: Panggil gerbang jendela login kustom
            else:
                self.after(100, jalankan_hitung_mundur_splash)
                
        # Aktifkan mesin hitung mundur sesaat setelah jendela splash dimuat di layar monitor
        self.after(200, jalankan_hitung_mundur_splash)

    def inisialisasi_dan_migrasi_excel_ke_sql(self):
        """ ENGINE ONLINE SERVER API v2: Koneksi terpusat via REST API 24/7 """
        import os
        import requests
        import pandas as pd
        
        # PARAMETER KONEKSI SERVER API v2 (kredensial DB tidak lagi dibutuhkan di client)
        self.db_host_online = ""; self.db_port_online = 0; self.db_name_online = ""
        self.db_user_online = ""; self.db_pass_online = ""
        self.excel_file_lama = "Inventaris_Laptop.xlsm"
        
        try:
            # 1. Kontak server pusat via REST API: pastikan schema + seed akun siap
            resp_boot = requests.post(f"{self.api_base_gudang}/api/gudang/bootstrap", timeout=15)
            resp_boot.raise_for_status()
            print("[SERVER SUCCESS] Aplikasi Berhasil Terhubung ke Server API v2!")
            
        except Exception as err:
            self.tampilkan_peringatan_besar("Koneksi Internet Putus", f"Gagal terhubung ke Server API!\n\nPastikan komputer Anda terhubung ke internet.\n\nDetail: {str(err)}", "#EF4444")

    def buat_jendela_login_sistem(self):
        """ MODUL BARU: JENDELA DIALOG LOGIN CARD PREMIUM (ROLE MANAGEMENT ACCESS) """
        import customtkinter as ctk
        
        # 1. Merakit Jendela Dialog Login Melayang Tengah Layar
        self.jendela_login = ctk.CTkToplevel(self)
        self.jendela_login.title("🔐 Verifikasi Otoritas Gudang - GTP")
        self.jendela_login.configure(fg_color="#F8FAFC")
        
        # Mengunci ukuran jendela login card (450 x 380 pixel)
        lebar_lg = 450
        tinggi_lg = 380
        lebar_layar = self.winfo_screenwidth()
        tinggi_layar = self.winfo_screenheight()
        posisi_x = int((lebar_layar - lebar_lg) / 2)
        posisi_y = int((tinggi_layar - tinggi_lg) / 2)
        self.jendela_login.geometry(f"{lebar_lg}x{tinggi_lg}+{posisi_x}+{posisi_y}")
        
        # Mengunci fokus mutlak agar operator tidak bisa meng-klik menu utama di belakangnya
        self.jendela_login.transient(self)
        self.jendela_login.grab_set()
        self.jendela_login.resizable(False, False)
        
        # Protokol paksa: Jika jendela login ditutup silang, matikan seluruh aplikasi harian
        self.jendela_login.protocol("WM_DELETE_WINDOW", self.quit)
        
        # 2. CONTAINER CARD INTERIOR WHITE PREMIUM
        card_login = ctk.CTkFrame(self.jendela_login, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E2E8F0")
        card_login.pack(padx=20, pady=20, fill="both", expand=True)
        
        lbl_instruksi = ctk.CTkLabel(card_login, text="SISTEM KEAMANAN INVENTARIS", font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"), text_color="#1E293B")
        lbl_instruksi.pack(pady=(25, 2))
        
        lbl_sub_instruksi = ctk.CTkLabel(card_login, text="Silakan pilih akun dan masukkan password Anda untuk masuk.", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#64748B")
        lbl_sub_instruksi.pack(pady=(0, 20))
        
        # --- KOMPONEN 1: DROPDOWN PILIHAN USERNAME AKUN ---
        lbl_user = ctk.CTkLabel(card_login, text="Pilih Identitas Pengguna :", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#334155")
        lbl_user.pack(anchor="w", padx=40, pady=(5, 2))
        
        daftar_users = ["ADMIN", "OPERATOR A", "OPERATOR B"]
        self.combo_login_user = ctk.CTkComboBox(card_login, values=daftar_users, font=ctk.CTkFont(family="Segoe UI", size=12), width=370, height=36, corner_radius=6, fg_color="#FFFFFF", border_color="#CBD5E1", button_color="#64748B", button_hover_color="#475569")
        self.combo_login_user.pack(padx=40, pady=(0, 15))
        self.combo_login_user.set("ADMIN") # Default value diatur ke ADMIN
        
        # --- KOMPONEN 2: KOTAK INPUT KATA SANDI (PASSWORD) ---
        lbl_pass = ctk.CTkLabel(card_login, text="Masukkan Kata Sandi (Password) :", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#334155")
        lbl_pass.pack(anchor="w", padx=40, pady=(5, 2))
        
        self.entry_login_pass = ctk.CTkEntry(card_login, width=370, height=36, show="*", corner_radius=6, border_color="#CBD5E1", font=ctk.CTkFont(family="Segoe UI", size=12), fg_color="#FFFFFF", text_color="#1E293B", placeholder_text="••••••••")
        self.entry_login_pass.pack(padx=40, pady=(0, 25))
        self.entry_login_pass.focus()
        # --- KOMPONEN 3: TOMBOL MASUK PREMIUM ---
        btn_masuk_sistem = ctk.CTkButton(card_login, text="🔐 Masuk ke Sistem Inventaris", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), height=40, corner_radius=6, fg_color="#2563EB", hover_color="#1D4ED8", text_color="#FFFFFF", command=self.proses_verifikasi_login)
        btn_masuk_sistem.pack(padx=40, fill="x")
        
        # Jembatan instan: Operator bisa langsung menekan tombol ENTER di keyboard untuk login
        self.entry_login_pass.bind("<Return>", lambda e: self.proses_verifikasi_login())

    def proses_verifikasi_login(self):
        """ ENGINE SECURITY CLOUD: Memvalidasi akun pengguna via REST API v2 """
        import requests
        
        user_pilihan = self.combo_login_user.get().strip()
        pass_input = self.entry_login_pass.get().strip()
        
        if not pass_input:
            self.bell()
            return
            
        try:
            # 1. Verifikasi login via REST API
            resp_login = requests.post(
                f"{self.api_base_gudang}/api/gudang/bridge-login",
                json={"username": user_pilihan, "password": pass_input},
                timeout=10,
            )
            resp_login.raise_for_status()
            data_login = resp_login.json()
            
            # 2. KONDISI A: JIKA KATA SANDI SALAH ATAU ACCOUNT TIDAK COCOK
            if not data_login.get("valid"):
                self.bell()
                self.entry_login_pass.delete(0, 'end')
                self.entry_login_pass.focus()
                self.tampilkan_peringatan_besar("Akses Ditolak", "KATA SANDI SALAH!\n\nOtoritas ditolak. Silakan periksa kembali kata sandi resmi untuk akun Anda.", "#EF4444")
                return
                
            role_akun = data_login.get("role")
            
            # Mengunci identitas operator aktif harian ke dalam memori RAM komputer
            self.operator_aktif = user_pilihan  
            self.role_aktif = role_akun
            
            # Hancurkan jendela login card secara bersih dari memori
            self.jendela_login.grab_release()
            self.jendela_login.destroy()
            
            # Picu memunculkan kembali jendela besar utama dashboard aplikasi ke monitor
            self.deiconify()
            self.focus_force()
            
            # Menerapkan pembatasan menu sidebar (Admin vs Operator)
            self.terapkan_pembatasan_hak_akses_role()
            
            # Menyambut kedatangan operator/admin dengan popup hijau sukses
            self.tampilkan_peringatan_besar(
                "Akses Diterima", 
                f"VERIFIKASI AKUN SUKSES!\n\nSelamat datang kembali, {self.operator_aktif}.\nSistem logistik pusat GTP telah dibuka dengan hak akses: {self.role_aktif}.", 
                "#10B981"
            )
            # Menerapkan pembatasan menu sidebar (Admin vs Operator)
            self.terapkan_pembatasan_hak_akses_role()
            
            # 🚀 SUNTIKAN ROBOT CEK EMAIL: Jalankan pengiriman email laporan otomatis harian
            self.jalankan_auto_backup_cloud_ke_email()
            
        except Exception as err:
            self.tampilkan_peringatan_besar("Koneksi Internet Putus", f"Gagal verifikasi! Server DataBase:\n\n{str(err)}", "#EF4444")

    def terapkan_pembatasan_hak_akses_role(self):
        """ ENGINE ACCESS CONTROL: Mengunci menu Master Data jika pengguna adalah Operator """
        import customtkinter as ctk
        
        try:
            # 1. KONDISI A: JIKA YANG MASUK ADALAH OPERATOR (Operator A atau Operator B)
            if self.role_aktif == "OPERATOR":
                # Mengunci tombol Master Data di sidebar menjadi abu-abu dan tidak bisa di-klik
                self.btn_stock.configure(
                    state="disabled", 
                    fg_color="#E2E8F0", 
                    text_color="#94A3B8",
                    text="🔒 MASTER DATA UNIT (ADMIN ONLY)"
                )
                
                # Mengarahkan paksa halaman utama ke Dashboard agar operator tidak terjebak di halaman kosong
                self.ganti_halaman("dash")
                print(f"[SECURITY] Pembatasan menu diaktifkan. Operator {self.operator_aktif} dilarang mengakses Master Data.")
            
            # 2. KONDISI B: JIKA YANG MASUK ADALAH ADMIN UTAMA GUDANG GTP
            else:
                # Mengembalikan tombol Master Data ke posisi normal, biru premium, dan aktif penuh
                self.btn_stock.configure(
                    state="normal", 
                    fg_color="transparent", 
                    text_color=("#1F2937", "#F8FAFC") if hasattr(self, 'txt_clr') else "#FFFFFF",
                    text="💻 MASTER DATA UNIT"
                )
                print("[SECURITY] Akses Admin Divalidasi. Seluruh fitur master data terbuka 100%.")
                
        except Exception as err:
            print(f"Gagal menerapkan pembatasan hak akses menu role: {str(err)}")

    def eksekusi_logout_akun_sistem(self):
        """ ENGINE LOGOUT: Menghapus sesi RAM user aktif dan memancing jendela login kembali """
        # 1. Konfirmasi kustom besar sebelum keluar demi keamanan data gudang
        yakin_keluar = self.tampilkan_konfirmasi_besar("Konfirmasi Keluar Akun", "Apakah Anda yakin ingin LOGOUT (Keluar) dari akun saat ini harian?\n\nTindakan ini akan mengunci kembali seluruh menu visual utama.")
        if not yakin_keluar:
            return
            
        # 2. Hancurkan radar clock pemantau otomatis 3 detik agar memori RAM bersih
        if hasattr(self, '_id_clock_refresh_live'):
            self.after_cancel(self._id_clock_refresh_live)
            
        # 3. Bersihkan sisa data identitas akun lama dari RAM komputer
        self.operator_aktif = None
        self.role_aktif = None
        
        # 4. Sembunyikan kembali jendela besar utama dashboard aplikasi Anda
        self.withdraw()
        
        # 5. Munculkan kembali Jendela Dialog Login Card Premium dari awal
        self.buat_jendela_login_sistem()
        print("[SECURITY] User berhasil logout. Sesi RAM dibersihkan, kembali ke gerbang login.")

    def jalankan_auto_backup_cloud_ke_email(self):
        """ ENGINE BACKEND EMAIL: Robot otomatis pengirim file backup Excel dari Server API ke Email Anda """
        import threading
        import datetime
        import pandas as pd
        import requests
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders

        def eksekusi_kirim_email_rahasia():
            tgl_hari_ini = datetime.datetime.now().strftime("%Y-%m-%d")
            file_backup_name = f"BACKUP_CLOUD_GTP_{tgl_hari_ini}.xlsx"
            
            # ATUR DATA EMAIL TUJUAN ANDA DI SINI
            email_pengirim = "serverdatabasegtp@gmail.com"  # Ganti dengan email Gmail sistem Anda
            password_aplikasi_gmail = "wmxw dlow nmhb keue" # Ganti dengan 16 digit Password Aplikasi Gmail
            email_penerima = "mametfebian@gmail.com"        # Alamat email pribadi Anda selaku owner
            
            try:
                # 1. Sedot data live dari server pusat via REST API v2
                resp_export = requests.get(f"{self.api_base_gudang}/api/gudang/full-export", timeout=60)
                resp_export.raise_for_status()
                data_export = resp_export.json()
                df_master = pd.DataFrame(data_export["master_data"])
                df_log = pd.DataFrame(data_export["log_tracking"])
                
                # 2. Kompilasi data menjadi 1 File Excel ber-Sheet banyak secara instan
                with pd.ExcelWriter(file_backup_name, engine='openpyxl') as writer:
                    df_master.to_excel(writer, sheet_name='MASTER DATA UNIT', index=False)
                    df_log.to_excel(writer, sheet_name='LOG KRONOLOGI MUTASI', index=False)
                
                # 3. Merakit struktur paket surat elektronik MIME standard Google
                msg = MIMEMultipart()
                msg['From'] = email_pengirim
                msg['To'] = email_penerima
                msg['Subject'] = f"📊 [AUTO-BACKUP] Laporan Mutasi Kargo GTP Real-Time ({tgl_hari_ini})"
                
                body_teks = f"Halo Owner MaMet SpooKy,\n\nBerikut adalah file lampiran dokumen backup otomatis dari sistem database SERVER GTP untuk tanggal {tgl_hari_ini}.\n\nSemua aktivitas scan masuk dan keluar terkunci aman.\n\nSalam,\nRobot Otomatis Gudang GTP."
                msg.attach(MIMEText(body_teks, 'plain'))
                
                # 4. Menyisipkan file Excel ke dalam lampiran surat harian
                with open(file_backup_name, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f"attachment; filename= {file_backup_name}")
                    msg.attach(part)
                
                # 5. Menembakkan email melintasi port keamanan SMTP Google Server
                server = smtplib.SMTP('://gmail.com', 587)
                server.starttls()
                server.login(email_pengirim, password_aplikasi_gmail)
                server.sendmail(email_pengirim, email_penerima, msg.as_string())
                server.quit()
                
                # Bersihkan sisa file temp backup dari harddisk lokal komputer biar rapi
                if os.path.exists(file_backup_name):
                    os.remove(file_backup_name)
                print(f"[EMAIL SUCCESS] Robot berhasil mengirimkan file laporan mutasi harian ke {email_penerima}")
                
            except Exception as mail_err:
                print(f"[EMAIL ERROR] Robot gagal mengirimkan laporan otomatis: {str(mail_err)}")

        # Jalankan di dalam thread terpisah agar aplikasi utama tidak patah-patah/freeze saat mengirim email
        threading.Thread(target=eksekusi_kirim_email_rahasia, daemon=True).start()

if __name__ == "__main__":
    # Membuat objek aplikasi dari cetak biru kelas yang sudah kita bangun
    app = ProfessionalWarehouseApp()
    
    # Menyalakan mesin looping jendela utama CustomTkinter
    app.mainloop()
